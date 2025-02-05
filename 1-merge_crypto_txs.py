# This script combines cryptocurrency transactions from multiple exchanges into one master file
# It handles transactions from Coinbase, Coinbase Pro, Kraken, Strike, and CashApp

# CryptoCompare API key - https://min-api.cryptocompare.com
# Optional: Script will use daily historical prices if no API key is provided
# With API key: Script can fetch minute-level historical prices (last 7 days only)
CRYPTOCOMPARE_API_KEY = 'YOUR-API-KEY-HERE'

# Import required Python libraries
import pandas as pd
import glob
from datetime import datetime, timedelta
import requests
import time
from decimal import Decimal, ROUND_HALF_UP
import json
from pathlib import Path
from typing import Dict, List
import logging
import openpyxl
import openpyxl.utils
import openpyxl.styles
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter

# Important date when LUNA became LUNC (Luna Classic)
LUNA_TRANSITION_DATE = datetime(2022, 5, 28)

# Add these constants near the top of the file with other style definitions
CALCULATED_FILL = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Light yellow
HISTORICAL_FILL = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')  # Light red

# Add these columns after the required_columns definition in merge_all_transactions()
calculated_cells = {
    'Subtotal': set(),  # Will store transaction IDs where Subtotal was calculated
    'Fee': set(),       # Will store transaction IDs where Fee was calculated
    'Total USD': set(), # Will store transaction IDs where Total USD was calculated
    'Spot Price': set() # Will store transaction IDs where Spot Price was calculated
}

# Set up logging to only create the log file if an error is actually logged
class ErrorOnlyFileHandler(logging.FileHandler):
    # This special file handler waits until an actual error occurs before creating the log file
    def __init__(self, filename, mode='a', encoding=None, delay=True):  # Set delay=True
        super().__init__(filename, mode, encoding, delay)
        self.error_occurred = False

    def emit(self, record):
        if record.levelno >= logging.ERROR:  # Only handle ERROR or higher
            self.error_occurred = True
            super().emit(record)

handler = ErrorOnlyFileHandler('ALL-MASTER-crypto-transactions.log')
handler.setLevel(logging.ERROR)
handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

logger = logging.getLogger()
logger.setLevel(logging.ERROR)
logger.addHandler(handler)

def log_error(message: str, print_to_screen: bool = True) -> None:
    # Helper function to handle error logging
    logging.getLogger().error(message)
    if print_to_screen:
        print(f"ERROR: {message}")

def load_price_cache():
    # Load previously fetched cryptocurrency prices from the "price_cache.json" cache file.
    # Returns an empty dictionary if no cache exists or if there's an error reading it.
    cache_file = Path('price_cache.json')
    if cache_file.exists():
        try:
            with open(cache_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading price cache: {e}")
    return {}

def save_price_cache(cache):
    # Save fetched cryptocurrency prices to the "price_cache.json" cache file.
    try:
        with open('price_cache.json', 'w') as f:
            json.dump(cache, f)
    except Exception as e:
        print(f"Error saving price cache: {e}")

# Initialize price cache
price_cache = load_price_cache()

def safe_concat(dfs: list, columns: list, **kwargs) -> pd.DataFrame:
    # Combine multiple data files while ensuring all required columns exist
    # Handles missing columns by adding them with appropriate default values
    if not dfs:
        return pd.DataFrame(columns=columns)

    
    # Ensure all DataFrames have all columns
    normalized_dfs = []
    for df in dfs:
        if isinstance(df, dict):
            df = pd.DataFrame([df])
        elif isinstance(df, pd.Series):
            df = pd.DataFrame([df.to_dict()])
        elif isinstance(df, list):
            df = pd.DataFrame(df)
            
        if isinstance(df, pd.DataFrame) and not df.empty:
            # Create a new DataFrame with the desired columns
            new_df = pd.DataFrame(columns=columns)
            
            # Copy over existing columns
            for col in columns:
                if col in df.columns:
                    new_df[col] = df[col]
                else:
                    # Initialize missing columns with empty strings for object dtype
                    # and 0 for numeric columns
                    if col in ['Amount', 'Subtotal', 'Fee', 'Total USD', 'Spot Price']:
                        new_df[col] = 0
                    else:
                        new_df[col] = ''
            
            normalized_dfs.append(new_df)
    
    if not normalized_dfs:
        return pd.DataFrame(columns=columns)
        
    return pd.concat(normalized_dfs, **kwargs)

def load_coinbase_transactions():
    # Load and process Coinbase transactions, handling special cases like:
    # - Convert transactions (split into buy/sell pairs)
    # - Staking rewards (renamed from Inflation Reward)
    # - Learning rewards and other income types
    try:
        # Find Coinbase transaction files
        files = glob.glob("*.xlsx") + glob.glob("*.csv")
        transaction_patterns = ['transactions', 'txs', 'transaction', 'tx', 'trans', 'action', 'activity', 'log', 'report', 'history']
        
        coinbase_files = [
            f for f in files 
            if 'coinbase' in f.lower() 
            and any(pattern in f.lower() for pattern in transaction_patterns)
            and 'pro' not in f.lower()  # Exclude Coinbase Pro files
        ]
        
        if not coinbase_files:
            print("Warning: No Coinbase transaction files found. Skipping Coinbase transactions.")
            return pd.DataFrame()
        
        print(f"  **  Loading Coinbase transactions...\n")
        print(f" {coinbase_files[0]}")
        df = read_transaction_file(coinbase_files[0])
        
        if df.empty:
            return df
            
        # Skip the header row if it's just "Timestamp"
        if df['Timestamp'].iloc[0] == 'Timestamp':
            df = df.iloc[1:]
            
        # Now continue with the normal processing
        df['Source'] = 'Coinbase'
        
        # Handle timestamp parsing with error checking
        def parse_timestamp(ts):
            try:
                # First try direct parsing
                return pd.to_datetime(ts)
            except:
                try:
                    # Try manual parsing if needed
                    return datetime.strptime(str(ts), '%Y-%m-%dT%H:%M:%SZ')
                except:
                    print(f"Warning: Could not parse timestamp: {ts}")
                    return None
                    
        df['Timestamp'] = df['Timestamp'].apply(parse_timestamp)
        df['Timestamp'] = df['Timestamp'].dt.tz_localize(None)
        
        # Clean and convert both columns to numeric, handling currency symbols
        subtotal = pd.to_numeric(
            df['Subtotal'].replace(r'[\$,]', '', regex=True),
            errors='coerce'
        )
        total = pd.to_numeric(
            df['Total (inclusive of fees and/or spread)'].replace(r'[\$,]', '', regex=True),
            errors='coerce'
        )
        
        # Create a mask for where we should use the Total instead of Subtotal
        use_total_mask = (
            subtotal.notna() & 
            total.notna() & 
            (abs(total) < abs(subtotal))
        )
        
        # Where the mask is True, use Total instead of Subtotal
        df.loc[use_total_mask, 'Subtotal'] = df.loc[use_total_mask, 'Total (inclusive of fees and/or spread)']
        
        # Now continue with the normal processing
        df['Source'] = 'Coinbase'
        df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.tz_localize(None)
        df = df.rename(columns={
            'Transaction Type': 'Type',
            'Quantity Transacted': 'Amount',
            'Subtotal': 'Subtotal',
            'Fees and/or Spread': 'Fee',
            'Price at Transaction': 'Spot Price'
        })
        
        # Clean monetary values using Decimal for precision
        df['Amount'] = df['Amount'].apply(lambda x: Decimal(str(x)) if pd.notnull(x) else Decimal('0'))
        df['Subtotal'] = df['Subtotal'].replace(r'[\$,]', '', regex=True).apply(lambda x: Decimal(str(x)) if pd.notnull(x) else Decimal('0'))
        df['Fee'] = df['Fee'].replace(r'[\$,]', '', regex=True).apply(lambda x: Decimal(str(x)) if pd.notnull(x) else Decimal('0'))
        df['Spot Price'] = df['Spot Price'].replace(r'[\$,]', '', regex=True).apply(lambda x: Decimal(str(x)) if pd.notnull(x) else Decimal('0'))
        
        # Calculate Total USD before processing Convert transactions
        df['Total USD'] = df.apply(
            lambda row: row['Subtotal'] + (abs(row['Fee']) * (1 if row['Subtotal'] >= 0 else -1)),
            axis=1
        )
        
        # Handle Convert transactions by creating two rows for each
        new_rows = []
        
        for idx, row in df.iterrows():
            if row['Type'] == 'Convert':
                # Parse the Notes field
                # Example: "Converted 21.40733 ALGO to 0.00022474 BTC"
                parts = row['Notes'].split(' ')
                sell_amount = Decimal(str(parts[1]))
                sell_asset = parts[2]
                buy_amount = Decimal(str(parts[4]))
                buy_asset = parts[5]
                
                # Get USD values directly from input file
                total_with_fees = abs(Decimal(str(row['Total (inclusive of fees and/or spread)']).replace('$', '').replace(',', '')))
                subtotal = abs(Decimal(str(row['Subtotal']).replace('$', '').replace(',', '')))
                fee = total_with_fees - subtotal  # Calculate fee as difference
                
                # Update the original row to be the sell transaction
                df.at[idx, 'Type'] = 'Sell'
                df.at[idx, 'Asset'] = sell_asset
                df.at[idx, 'Amount'] = sell_amount
                df.at[idx, 'Fee'] = fee
                df.at[idx, 'Total USD'] = total_with_fees  # Use the total with fees value
                df.at[idx, 'Spot Price'] = total_with_fees / sell_amount
                
                # Create the buy transaction as a new row
                buy_row = row.copy()
                buy_row['Type'] = 'Buy'
                buy_row['Asset'] = buy_asset
                buy_row['Amount'] = buy_amount
                buy_row['Fee'] = Decimal('0')
                buy_row['Total USD'] = total_with_fees  # Use same total with fees value
                buy_row['Spot Price'] = total_with_fees / buy_amount
                new_rows.append(buy_row)
        
        # Add the new buy rows to the DataFrame
        if new_rows:
            df = safe_concat([df, pd.DataFrame(new_rows)], columns=df.columns, ignore_index=True)
        
        # Rename "Inflation Reward" to "Staking Income"
        df.loc[df['Type'] == 'Inflation Reward', 'Type'] = 'Staking Income'
        
        # Ensure numeric columns are properly typed
        numeric_columns = ['Amount', 'Total USD', 'Fee', 'Spot Price']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: Decimal(str(x)) if pd.notnull(x) else Decimal('0'))
        
        return df[['ID', 'Timestamp', 'Source', 'Type', 'Asset', 'Amount', 'Subtotal', 'Fee', 'Total USD', 'Spot Price', 'Notes']]
    except Exception as e:
        print(f"Error loading Coinbase transactions: {str(e)}")
        return pd.DataFrame()

def load_coinbase_pro_transactions():
    # Load and process Coinbase Pro trades, matching fills with their orders
    # Handles both buy and sell transactions from the fills report
    try:
        dfs = []
        files = glob.glob("*.xlsx") + glob.glob("*.csv")
        
        # Match any combination of fills/fill and coinbasepro/coinbase-pro
        fill_patterns = ['fill', 'fills']
        pro_patterns = ['coinbasepro', 'coinbase-pro']
        pro_files = [
            f for f in files 
            if any(pro in f.lower() for pro in pro_patterns)
            and any(fill in f.lower() for fill in fill_patterns)
        ]
        
        if not pro_files:
            print("Warning: No Coinbase Pro files found. Skipping Coinbase Pro transactions.")
            return pd.DataFrame()
        
        print(f"  **  Loading Coinbase Pro transactions...\n")
        for file in pro_files:
            print(f" {file}")
            df = read_transaction_file(file)
            
            if df.empty:
                continue
            
            if 'portfolio' in df.columns:
                df = df.drop('portfolio', axis=1)
            
            # Add Notes column if it doesn't exist
            if 'Notes' not in df.columns:
                df['Notes'] = ''
            
            df['Source'] = 'Coinbase Pro'
            df['Timestamp'] = pd.to_datetime(df['created at']).dt.tz_localize(None)
            df['Asset'] = df['product'].str.split('-').str[0]
            df['Type'] = df['side']
            
            # Rename columns and convert ID to string
            df = df.rename(columns={
                'trade id': 'ID',
                'size': 'Amount',
                'total': 'Total USD',
                'fee': 'Fee',
                'price': 'Spot Price'
            })
            df['ID'] = df['ID'].astype(str)  # Convert ID to string
            
            # Calculate Subtotal
            df['Subtotal'] = df.apply(
                lambda row: (abs(row['Total USD']) - (abs(row['Fee']))),
                axis=1
            )
            
            dfs.append(df[['ID', 'Timestamp', 'Source', 'Type', 'Asset', 'Amount', 'Subtotal', 'Fee', 'Total USD', 'Spot Price', 'Notes']])
        
        return pd.concat(dfs) if dfs else pd.DataFrame()
        
    except Exception as e:
        print(f"Error loading Coinbase Pro transactions: {str(e)}")
        return pd.DataFrame()

def load_kraken_transactions():
    # Load and combine Kraken ledger entries with trade history
    # Matches trades with their orders and handles various transaction types
    try:
        files = glob.glob("*.xlsx") + glob.glob("*.csv")
        
        # Find ledger file
        ledger_patterns = ['kraken-ledgers', 'kraken_ledgers', 'kraken-ledger', 'kraken_ledger']
        ledger_files = [f for f in files if any(pattern in f.lower() for pattern in ledger_patterns)]
        
        if not ledger_files:
            print("Warning: No Kraken ledger file found. Skipping Kraken transactions.")
            return pd.DataFrame()
            
        ledger_file = ledger_files[0]
        print(f"  **  Loading Kraken ledger...\n")
        print(f" {ledger_file}")
        ledger_df = read_transaction_file(ledger_file).reset_index(drop=True)
        
        # Find trades file
        trades_patterns = ['kraken-trades', 'kraken_trades', 'kraken-trade', 'kraken_trade']
        trades_files = [f for f in files if any(pattern in f.lower() for pattern in trades_patterns)]
        
        trades_df = pd.DataFrame()
        if not trades_files:
            print("Warning: No Kraken trades file found. Proceeding with ledger only.")
        elif trades_files:
            trades_file = trades_files[0]
            print(f"\n  **  Loading Kraken trades...\n")
            print(f" {trades_file}")
            trades_df = read_transaction_file(trades_file).reset_index(drop=True)
        
        # Create a mapping of txid to ordertxid from trades file
        ordertxid_map = {}
        if not trades_df.empty:
            for _, row in trades_df.iterrows():
                ordertxid_map[row['txid']] = row['ordertxid']
        
        # Create price mapping from trades file
        price_map = {}
        if not trades_df.empty:
            for _, row in trades_df.iterrows():
                # Only map prices for crypto assets traded against USD
                if 'pair' in row and '/USD' in row['pair'] and row['txid']:
                    base_asset = row['pair'].split('/')[0]  # Get the crypto asset part
                    price_map[row['txid']] = {
                        'price': row['price'],
                        'asset': base_asset
                    }
        
        # Process ledger entries
        ledger_df['Timestamp'] = pd.to_datetime(ledger_df['time']).dt.tz_localize(None)
        ledger_df['Source'] = 'Kraken'
        ledger_df['ID'] = ledger_df['txid']
        
        # Remove .S from asset names
        ledger_df['asset'] = ledger_df['asset'].str.split('.').str[0]
        
        # Create empty DataFrame for results with correct columns
        result_columns = ['ID', 'Timestamp', 'Source', 'Type', 'Asset', 'Amount', 'Subtotal', 'Fee', 'Total USD', 'Spot Price', 'Notes']
        all_transactions = pd.DataFrame(columns=result_columns)
        
        # Process trades
        if not ledger_df.empty:
            trades = ledger_df[ledger_df['type'] == 'trade'].copy()
            trades = trades.sort_index()  # Ensure original file order is preserved
            trades['original_index'] = trades.index  # Add original index column
            
            trades_transactions = pd.DataFrame(columns=result_columns + ['original_index'])
            
            for refid, pair in trades.groupby('refid', sort=False):
                if len(pair) == 2:
                    # Get the rows in original file order - Kraken always lists SELL first
                    sell_row = pair.iloc[0]  # First row is always the SELL
                    buy_row = pair.iloc[1]   # Second row is always the BUY
                    
                    # Verify this assumption - if not true, we need to swap
                    if sell_row['amount'] > 0:  # If first row is actually a buy
                        sell_row, buy_row = buy_row, sell_row  # swap them
                    
                    # Get ordertxid and set up notes
                    ordertxid = ordertxid_map.get(refid, '')
                    notes = f"Order: {ordertxid}" if ordertxid else ""
                    notes += f" ~~ Trade: {refid}" if notes else f"Trade: {refid}"
                    
                    # Check if this is a USD pair trade
                    is_usd_pair = 'USD' in [sell_row['asset'], buy_row['asset']]
                    
                    if is_usd_pair:
                        # For USD pairs, use the actual USD values directly
                        if sell_row['asset'] == 'USD':
                            # USD is being sold (buying crypto)
                            total_usd = abs(sell_row['amount'])  # Total USD is what's being sold
                            spot_price = total_usd / abs(buy_row['amount'])
                            fee_in_usd = abs(buy_row['fee']) * spot_price if buy_row['fee'] != 0 else abs(sell_row['fee'])
                            base_subtotal = total_usd - fee_in_usd  # Subtotal is Total minus fee
                            
                            # For USD pairs, both sides share the same fee
                            sell_fee_in_usd = fee_in_usd
                            buy_fee_in_usd = fee_in_usd
                        else:
                            # USD is being bought (selling crypto)
                            total_usd = abs(buy_row['amount'])  # Total USD is what's being bought
                            spot_price = total_usd / abs(sell_row['amount'])
                            fee_in_usd = abs(sell_row['fee']) * spot_price if sell_row['fee'] != 0 else abs(buy_row['fee'])
                            base_subtotal = total_usd - fee_in_usd  # Subtotal is Total minus fee
                            
                            # For USD pairs, both sides share the same fee
                            sell_fee_in_usd = fee_in_usd
                            buy_fee_in_usd = fee_in_usd
                    else:
                        # For non-USD pairs, calculate fees separately for each side
                        base_row = sell_row if abs(sell_row['amountusd']) > abs(buy_row['amountusd']) else buy_row
                        total_usd = abs(base_row['amountusd'])  # amountusd is pre-fee
                        
                        # Calculate spot prices for both sides
                        sell_spot_price = total_usd / abs(sell_row['amount'])
                        buy_spot_price = total_usd / abs(buy_row['amount'])
                        
                        # Calculate fees separately for each side
                        sell_fee_in_usd = abs(sell_row['fee']) * sell_spot_price if sell_row['fee'] != 0 else 0
                        buy_fee_in_usd = abs(buy_row['fee']) * buy_spot_price if buy_row['fee'] != 0 else 0
                        
                        # For non-USD pairs, subtotal is total minus all fees to balance both sides
                        base_subtotal = total_usd - sell_fee_in_usd - buy_fee_in_usd

                    transactions = [
                        # SELL transaction
                        {
                            'ID': sell_row['txid'],
                            'Timestamp': sell_row['Timestamp'],
                            'Source': 'Kraken',
                            'Type': 'Sell',
                            'Asset': sell_row['asset'],
                            'Amount': abs(sell_row['amount']),
                            'Subtotal': base_subtotal,  # Same subtotal for both sides
                            'Fee': sell_fee_in_usd,  # Use sell-side fee
                            'Total USD': total_usd,
                            'Spot Price': total_usd / abs(sell_row['amount']),
                            'Notes': f"{notes} ~~ Base Asset: {base_asset}" if not is_usd_pair else notes
                        },
                        # BUY transaction
                        {
                            'ID': buy_row['txid'],
                            'Timestamp': buy_row['Timestamp'],
                            'Source': 'Kraken',
                            'Type': 'Buy',
                            'Asset': buy_row['asset'],
                            'Amount': abs(buy_row['amount']),
                            'Subtotal': base_subtotal,  # Same subtotal for both sides
                            'Fee': buy_fee_in_usd,  # Use buy-side fee
                            'Total USD': total_usd,
                            'Spot Price': total_usd / abs(buy_row['amount']),
                            'Notes': f"{notes} ~~ Base Asset: {base_asset}" if not is_usd_pair else notes
                        }
                    ]
                    
                    # Add original_index to transactions
                    transactions[0]['original_index'] = pair.iloc[0]['original_index']
                    transactions[1]['original_index'] = pair.iloc[1]['original_index']
                    
                    trades_transactions = safe_concat([trades_transactions, transactions], 
                                                   columns=result_columns + ['original_index'],
                                                   ignore_index=False)
            
            # Sort and drop original_index only for trades
            if not trades_transactions.empty:
                trades_transactions = trades_transactions.sort_values('original_index')
                trades_transactions = trades_transactions.drop('original_index', axis=1)
                all_transactions = safe_concat([all_transactions, trades_transactions], 
                                            columns=result_columns,
                                            ignore_index=True)
        
        # Process non-trade entries
        non_trades = ledger_df[ledger_df['type'] != 'trade'].copy()
        
        # Filter transfers
        non_trades = non_trades[
            (non_trades['type'] != 'transfer') | 
            ((non_trades['type'] == 'transfer') & (non_trades['subtype'] == 'spotfromfutures'))
        ]
        
        # Rename special types
        non_trades.loc[
            (non_trades['type'] == 'transfer') & 
            (non_trades['subtype'] == 'spotfromfutures'), 
            'type'
        ] = 'Staking Income'
        
        non_trades.loc[non_trades['type'] == 'staking', 'type'] = 'Staking Income'
        
        if not non_trades.empty:
            non_trades_formatted = pd.DataFrame({
                'ID': non_trades['txid'],
                'Timestamp': non_trades['Timestamp'],
                'Source': 'Kraken',
                'Type': non_trades['type'],
                'Asset': non_trades['asset'],
                'Amount': non_trades['amount'].abs(),
                'Subtotal': non_trades['amountusd'].abs(),
                'Fee': non_trades['fee'],
                'Total USD': non_trades['amountusd'].abs() + non_trades['fee'],
                'Spot Price': 0,
                'Notes': ''
            })
            
            all_transactions = safe_concat([all_transactions, non_trades_formatted], 
                                        columns=result_columns,
                                        ignore_index=True)
        
        # When creating trade entries, add ordertxid to Notes
        if not trades_df.empty:
            for idx, row in trades_df.iterrows():
                if row['ordertxid']:
                    if 'Notes' not in trades_df.columns:
                        trades_df['Notes'] = ''
                    trades_df.loc[idx, 'Notes'] = f"Order: {row['ordertxid']} " + str(trades_df.loc[idx, 'Notes'])
        
        return all_transactions[result_columns]
        
    except Exception as e:
        print(f"Error loading Kraken transactions: {str(e)}")
        return pd.DataFrame(columns=['ID', 'Timestamp', 'Source', 'Type', 'Asset', 'Amount', 'Total USD', 'Fee', 'Spot Price', 'Notes'])

def load_strike_transactions():
    # Load and process Strike Bitcoin transactions
    # Handles buys, sells, and withdrawals from Strike's transaction report
    try:
        # Find Strike transaction files
        files = glob.glob("*.xlsx") + glob.glob("*.csv")
        transaction_patterns = ['transactions', 'txs', 'transaction', 'tx', 'trans', 'action', 'activity', 'log', 'report', 'history']

        strike_files = [
            f for f in files 
            if 'strike' in f.lower() 
            and any(pattern in f.lower() for pattern in transaction_patterns)
        ]
        
        if not strike_files:
            print("Warning: No Strike transaction files found. Skipping Strike transactions.")
            return pd.DataFrame()
            
        dfs = []
        
        print(f"  **  Loading Strike transactions...\n")
        for file in strike_files:
            print(f" {file}")
            df = read_transaction_file(file)
            
            if df.empty:
                continue
                
            # Filter rows based on Transaction Type and State
            df = df[
                (df['Transaction Type'].isin(['Trade', 'Withdrawal'])) & 
                (df['State'] == 'Completed')
            ]
            
            # Create timestamp from date and time columns
            df['Timestamp'] = pd.to_datetime(
                df['Completed Date (UTC)'] + ' ' + df['Completed Time (UTC)']
            ).dt.tz_localize(None)
            
            # Set Source
            df['Source'] = 'Strike'
            
            # Determine Transaction Type
            def get_transaction_type(row):
                if row['Transaction Type'] == 'Withdrawal':
                    return 'Send'
                elif (row['Transaction Type'] == 'Trade' and 
                      row['Amount 1'] < 0 and 
                      row['Currency 1'] == 'USD' and 
                      row['Currency 2'] == 'BTC'):
                    return 'Buy'
                elif (row['Transaction Type'] == 'Trade' and 
                      ((row['Currency 1'] == 'BTC' and row['Currency 2'] == 'USD') or
                       (row['Currency 1'] == 'USD' and row['Currency 2'] == 'BTC' and 
                        row['Amount 1'] > 0 and row['Amount 2'] < 0))):
                    return 'Sell'
                return 'Unknown'
            
            df['Type'] = df.apply(get_transaction_type, axis=1)
            
            # Determine Asset
            df['Asset'] = df.apply(
                lambda row: row['Currency 1'] if row['Currency 2'] == 'USD' else row['Currency 2'],
                axis=1
            )
            
            # Map other fields
            df['ID'] = df['Transaction ID']
            df['Amount'] = pd.to_numeric(df['Amount 2'], errors='coerce').fillna(0)
            df['Total USD'] = pd.to_numeric(df['Amount 1'], errors='coerce').fillna(0)
            
            # Handle fees - replace NaN with 0 and convert to numeric
            df['Fee 1'] = pd.to_numeric(df['Fee 1'], errors='coerce').fillna(0)
            df['Fee 2'] = pd.to_numeric(df['Fee 2'], errors='coerce').fillna(0)
            df['Fee'] = df['Fee 1'].abs() + df['Fee 2'].abs()
            
            # Handle Spot Price
            df['Spot Price'] = pd.to_numeric(df['BTC Price'], errors='coerce').fillna(0)
            
            # Calculate Subtotal
            df['Subtotal'] = df['Total USD'].abs() - df['Fee']
            
            # Add empty Notes column
            df['Notes'] = ''
            
            # Select and reorder columns
            result_df = df[[
                'ID', 'Timestamp', 'Source', 'Type', 'Asset', 'Amount',
                'Subtotal', 'Fee', 'Total USD', 'Spot Price', 'Notes'
            ]]
            
            dfs.append(result_df)
        
        if dfs:
            final_df = pd.concat(dfs, ignore_index=True)

            return final_df
        return pd.DataFrame()
        
    except Exception as e:
        print(f"Error loading Strike transactions: {str(e)}")
        return pd.DataFrame()

def load_cashapp_transactions():
    # Load and process CashApp Bitcoin transactions
    # Handles buys, sells, and transfers from CashApp's transaction report
    try:
        # Find CashApp transaction files
        files = glob.glob("*.xlsx") + glob.glob("*.csv")
        transaction_patterns = ['transactions', 'txs', 'transaction', 'tx', 'trans', 'action', 'activity', 'log', 'report', 'history']
        cashapp_files = [
            f for f in files 
            if any(f.lower().startswith(p) for p in ['cash_app', 'cashapp'])
            and any(pattern in f.lower() for pattern in transaction_patterns)
        ]
        
        if not cashapp_files:
            print("Warning: No CashApp files found. Skipping CashApp transactions.")
            return pd.DataFrame()
            
        dfs = []
        transaction_counter = 1  # For generating IDs
        
        for file in cashapp_files:
            print(f"  **  Loading CashApp transactions...\n")
            print(f" {file}")
            df = read_transaction_file(file)
            
            if df.empty:
                continue
                
            # Filter rows based on Transaction Type and Status
            df = df[
                (df['Transaction Type'].str.startswith('Bitcoin', na=False)) & 
                (df['Status'].fillna('') == 'COMPLETE')
            ]
            
            if df.empty:
                continue
                
            # Create timestamp from Date column - handle various timezone formats
            def parse_timestamp(date_str):
                try:
                    # Split into date/time and timezone parts
                    parts = date_str.rsplit(' ', 1)  # Split from right side once
                    if len(parts) == 2:
                        datetime_str, tz = parts
                        # Parse the datetime string
                        dt = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
                        
                        # Adjust for timezone
                        if tz == 'CDT':
                            dt = dt - timedelta(hours=5)  # CDT is UTC-5
                        elif tz == 'CST':
                            dt = dt - timedelta(hours=6)  # CST is UTC-6
                        
                        return dt
                    return pd.to_datetime(date_str)
                except Exception as e:
                    print(f"Error parsing date '{date_str}': {str(e)}")
                    return None

            # Convert dates and handle timezone
            df['Timestamp'] = df['Date'].apply(lambda x: parse_timestamp(str(x)))
            
            # Drop rows with invalid timestamps
            invalid_timestamps = df['Timestamp'].isna()
            if invalid_timestamps.any():
                print(f"Warning: Dropping {invalid_timestamps.sum()} rows with invalid timestamps")
                df = df.dropna(subset=['Timestamp'])
            
            # Set Source
            df['Source'] = 'CashApp'
            
            # Map Transaction Types
            type_mapping = {
                'Bitcoin Buy': 'Buy',
                'Bitcoin Sell': 'Sell',
                'Bitcoin Withdrawal': 'Send',
                'Bitcoin Send': 'Send'
            }
            df['Type'] = df['Transaction Type'].map(type_mapping)
            
            # Generate IDs for missing Transaction IDs
            next_id = transaction_counter
            def generate_id(x):
                nonlocal next_id
                if pd.notna(x):
                    return x
                else:
                    current_id = f'cashapp-{next_id:06d}'
                    next_id += 1
                    return current_id
                    
            df['ID'] = df['Transaction ID'].apply(generate_id)
            # Update counter for next file
            transaction_counter = next_id
            
            # Map other fields
            df['Asset'] = df['Asset Type']
            df['Spot Price'] = df['Asset Price'].apply(clean_price_string)
            df['Fee'] = df['Fee'].apply(clean_price_string).fillna(0).abs()
            
            # Create a mask for Buy/Sell transactions
            buy_sell_mask = df['Type'].isin(['Buy', 'Sell'])
            
            # Handle Amount based on transaction type
            df['Amount'] = df.apply(
                lambda row: pd.to_numeric(row['Asset Amount'], errors='coerce')
                if row['Type'] in ['Buy', 'Sell']
                else pd.to_numeric(row['Amount'], errors='coerce'),
                axis=1
            ).fillna(0)
            
            # Handle Total USD and Subtotal based on transaction type
            df['Total USD'] = df.apply(
                lambda row: clean_price_string(row['Net Amount'])  # Use Net Amount for Total USD
                if row['Type'] in ['Buy', 'Sell']
                else 0,
                axis=1
            ).fillna(0).abs()
            
            df['Subtotal'] = df.apply(
                lambda row: clean_price_string(row['Net Amount']) - clean_price_string(row['Fee'])  # Calculate Subtotal as Net Amount - Fee
                if row['Type'] in ['Buy', 'Sell']
                else 0,
                axis=1
            ).abs()
            
            # Copy Notes
            df['Notes'] = df['Notes'].fillna('')
            
            # Select and reorder columns
            result_df = df[[
                'ID', 'Timestamp', 'Source', 'Type', 'Asset', 'Amount',
                'Subtotal', 'Fee', 'Total USD', 'Spot Price', 'Notes'
            ]]
            
            dfs.append(result_df)
        
        if dfs:
            final_df = pd.concat(dfs, ignore_index=True)
            return final_df
        return pd.DataFrame()
        
    except Exception as e:
        print(f"Error loading CashApp transactions: {str(e)}")
        import traceback
        traceback.print_exc()  # This will print the full error traceback
        return pd.DataFrame()

def clean_price_string(value):
    # Clean up price strings to numeric values and removes $ signs and commas
    # Returns 0 if the value can't be converted to a number
    if pd.isna(value):
        return 0

    try:
        # If it's already a number, return it
        if isinstance(value, (int, float)):
            return float(value)
        # Clean string: remove $, commas, and whitespace
        cleaned = str(value).replace('$', '').replace(',', '').strip()
        return float(cleaned) if cleaned else 0
    except:
        return 0

def standardize_transaction_values(df: pd.DataFrame) -> pd.DataFrame:
    # Standardize transaction signs for consistency:
    # + (positive): Buys, rewards, staking income
    # - (negative): Sells, sends, withdrawals
    try:
        # Make a copy to avoid modifying the original

        df = df.copy()
        
        # Ensure all fees are positive
        df['Fee'] = df['Fee'].abs()
        
        # Convert Type to lowercase for consistent comparison
        df['Type_lower'] = df['Type'].str.lower()
        
        print("\n\nUnique transaction types:", df['Type_lower'].unique())
        
        # Filter out Exchange Withdrawal rows
        df = df[~df['Type_lower'].str.contains('exchange withdrawal', na=False)]
        
        # Define transaction type patterns
        sell_patterns = ['sell', 'send', 'convert']
        buy_patterns = ['buy', 'receive', 'reward', 'deposit', 'staking', 'dividend']
        
        # Create masks for different transaction types - with error checking
        def safe_pattern_check(x, patterns):
            try:
                if pd.isna(x):
                    return False
                return any(pattern in str(x).lower() for pattern in patterns)
            except Exception as e:
                print(f"Error in pattern check for value '{x}': {str(e)}")
                return False
        
        df['is_sell'] = df['Type_lower'].apply(lambda x: safe_pattern_check(x, sell_patterns))
        df['is_buy'] = df['Type_lower'].apply(lambda x: safe_pattern_check(x, buy_patterns))
        df['is_admin_debit'] = df['Type_lower'].str.contains('admin debit', na=False)
        df['is_withdrawal'] = df['Type_lower'].str.contains('withdrawal', na=False)
        
        print("\nTransaction count...")
        # Print mask counts for debugging
        print(f"   Sell transactions: {df['is_sell'].sum()}")
        print(f"   Buy transactions: {df['is_buy'].sum()}")
        print(f"   Admin debit transactions: {df['is_admin_debit'].sum()}")
        print(f"   Withdrawal transactions: {df['is_withdrawal'].sum()}")
        
        # Apply standardization using masks
        if df['is_sell'].any():
            df.loc[df['is_sell'], 'Amount'] = -df.loc[df['is_sell'], 'Amount'].abs()
            df.loc[df['is_sell'], 'Subtotal'] = -df.loc[df['is_sell'], 'Subtotal'].abs()
        
        if df['is_buy'].any():
            df.loc[df['is_buy'], 'Amount'] = df.loc[df['is_buy'], 'Amount'].abs()
            df.loc[df['is_buy'], 'Subtotal'] = df.loc[df['is_buy'], 'Subtotal'].abs()
        
        if df['is_admin_debit'].any():
            df.loc[df['is_admin_debit'], 'Amount'] = -df.loc[df['is_admin_debit'], 'Amount'].abs()
            df.loc[df['is_admin_debit'], 'Subtotal'] = -df.loc[df['is_admin_debit'], 'Subtotal'].abs()
        
        if df['is_withdrawal'].any():
            df.loc[df['is_withdrawal'], 'Amount'] = -df.loc[df['is_withdrawal'], 'Amount'].abs()
            df.loc[df['is_withdrawal'], 'Subtotal'] = 0
        
        print("\nFilling in the missing blanks...")
        # Calculate Total USD based on Subtotal and Fee
        df['Total USD'] = df.apply(
            lambda row: row['Subtotal'] + (abs(row['Fee']) * (1 if row['Subtotal'] >= 0 else -1)),
            axis=1
        )
        
        # Drop temporary columns
        df = df.drop(['Type_lower', 'is_sell', 'is_buy', 'is_admin_debit', 'is_withdrawal'], axis=1)
        
        return df
        
    except Exception as e:
        print(f"Error in standardize_transaction_values: {str(e)}")
        print(f"Error occurred at line: {e.__traceback__.tb_lineno}")
        raise

def process_kraken_groups(asset_df: pd.DataFrame) -> pd.DataFrame:
    # Group related Kraken transactions that share the same order ID
    # This combines multiple parts of the same trade into one transaction
    kraken_groups = {}
    non_kraken_rows = []
    
    for idx, row in asset_df.iterrows():
        if (row['Source'] == 'Kraken' and 'Notes' in row and 
            isinstance(row['Notes'], str) and 'Order:' in row['Notes']):
            ordertxid = row['Notes'].split('Order:')[1].split()[0]
            if ordertxid not in kraken_groups:
                kraken_groups[ordertxid] = []
            kraken_groups[ordertxid].append(row)
        else:
            non_kraken_rows.append(row)
    
    # Process Kraken groups first
    for ordertxid, group in kraken_groups.items():
        if len(group) > 1:
            # Sort group by type_order first
            group = sorted(group, key=lambda x: x['type_order'])
            # Verify these transactions should be merged
            first_tx = group[0]
            if all(should_merge_transactions(first_tx.to_dict(), tx.to_dict()) 
                  for tx in group[1:]):
                combined = combine_transactions(pd.DataFrame(group))
                non_kraken_rows.append(combined)
            else:
                non_kraken_rows.extend(group)
        else:
            non_kraken_rows.extend(group)
    
    # Convert back to DataFrame and sort
    return pd.DataFrame(non_kraken_rows).sort_values(['Timestamp', 'type_order'])

def assign_lot_ids_and_group(df: pd.DataFrame, time_window_seconds: int = 90) -> pd.DataFrame:
    #Group similar transactions by source & type within time windows, assign lot IDs chronologically.
    df = df.copy()
    df['Lot ID'] = ''
    

    # Define receiving types
    receiving_types = [
        'buy',
        'advanced trade buy',
        'receive',
        'learning reward',
        'staking income',
        'dividend',
        'convert'
    ]
    
    # First, handle all grouping
    result_dfs = []
    
    for asset in df['Asset'].unique():
        if asset == 'USD':
            continue
            
        asset_df = df[df['Asset'] == asset].copy()
        
        # Add type ordering for same-timestamp transactions
        asset_df['type_order'] = asset_df['Type'].str.lower().map({'sell': 0, 'buy': 1}).fillna(2)
        
        # Process Kraken transactions first
        asset_df = process_kraken_groups(asset_df)
        asset_df = asset_df.drop('type_order', axis=1)
        
        # Initialize variables for grouping
        current_group = []
        all_groups = []
        
        # Group transactions within time window
        for idx, row in asset_df.iterrows():
            if not current_group:
                current_group = [row]
            else:
                if should_merge_transactions(current_group[0].to_dict(), row.to_dict(), time_window_seconds):
                    current_group.append(row)
                else:
                    if len(current_group) > 1:
                        # Sort group by type before combining (sell first)
                        current_group = sorted(current_group, key=lambda x: 0 if x['Type'].lower() == 'sell' else 1)
                        all_groups.append(combine_transactions(pd.DataFrame(current_group)))
                    else:
                        all_groups.append(current_group[0])
                    current_group = [row]
        
        # Handle last group
        if current_group:
            if len(current_group) > 1:
                # Sort last group by type before combining (sell first)
                current_group = sorted(current_group, key=lambda x: 0 if x['Type'].lower() == 'sell' else 1)
                all_groups.append(combine_transactions(pd.DataFrame(current_group)))
            else:
                all_groups.append(current_group[0])
        
        if all_groups:
            result_dfs.append(pd.DataFrame(all_groups))
    
    # Combine all results
    if not result_dfs:
        return df
        
    # Combine and sort all transactions
    result = pd.concat(result_dfs, ignore_index=True)
    
    # Add type_order for final sorting
    result['type_order'] = result['Type'].str.lower().map({'sell': 0, 'buy': 1}).fillna(2)
    result = result.sort_values(['Timestamp', 'type_order']).reset_index(drop=True)
    result = result.drop('type_order', axis=1)
    
    # NOW assign lot IDs in a separate pass
    for asset in result['Asset'].unique():
        if asset == 'USD':
            continue
            
        # Get asset transactions
        asset_mask = result['Asset'] == asset
        
        # Assign lot IDs sequentially for receiving transactions only
        lot_counter = 1
        for idx in result[asset_mask].index:
            tx_type = result.loc[idx, 'Type'].lower()
            # Strict check: must be in receiving_types AND not contain 'sell'
            if tx_type in receiving_types and 'sell' not in tx_type:
                result.loc[idx, 'Lot ID'] = f"{asset}-{lot_counter:05d}"
                lot_counter += 1
    
    return result

def get_decimal_places(value_str: str) -> int:
    # Count how many decimal places are in a number
    # For example: "1.23" has 2 decimal places, "1.0" has 1
    try:
        value_str = str(value_str)

        if '.' in value_str:
            return len(value_str.split('.')[1])
        return 0
    except:
        return 8  # Default to 8 decimal places if we can't determine

def normalize_amount(amount, precision=None):
    # Convert numbers to exact decimal values with proper precision
    # Ensures we don't lose any decimal places during calculations
    if amount is None or pd.isna(amount):
        return Decimal('0')

    
    # Convert to string first to preserve original precision
    amount_str = str(amount)
    
    # Determine precision if not specified
    if precision is None:
        precision = get_decimal_places(amount_str)
    
    # Create Decimal with specific precision
    d = Decimal(amount_str)
    return d.quantize(Decimal('0.' + '0' * precision), rounding=ROUND_HALF_UP)

def combine_transactions(group: pd.DataFrame) -> pd.Series:
    # Merge multiple related transactions into one
    # Combines amounts, fees, and calculates weighted average prices
    # Also organizes notes for grouped transactions

    # Sort by timestamp to use earliest time
    group = group.sort_values('Timestamp')
    

    # Determine precision from the Amount column
    max_precision = max(
        get_decimal_places(str(amount))
        for amount in group['Amount']
        if pd.notna(amount)
    )
    
    # Convert amounts to Decimal and sum
    amount_sum = sum(
        normalize_amount(amount, max_precision)
        for amount in group['Amount']
        if pd.notna(amount)
    )
    
    # Calculate weighted average for spot price
    non_zero_mask = group['Spot Price'] > 0
    if non_zero_mask.any():
        valid_amounts = [normalize_amount(amt, max_precision) for amt in group.loc[non_zero_mask, 'Amount'].abs()]
        valid_prices = [normalize_amount(price, max_precision) for price in group.loc[non_zero_mask, 'Spot Price']]
        weighted_spot_price = sum(p * a for p, a in zip(valid_prices, valid_amounts)) / sum(valid_amounts)
    else:
        weighted_spot_price = Decimal('0')
    
    # Sum other numeric values with proper precision
    subtotal_sum = sum(normalize_amount(sub, max_precision) for sub in group['Subtotal'] if pd.notna(sub))
    fee_sum = sum(normalize_amount(fee, max_precision) for fee in group['Fee'] if pd.notna(fee))
    total_usd_sum = sum(normalize_amount(total, max_precision) for total in group['Total USD'] if pd.notna(total))
    
    # Organize Notes for Kraken transactions
    if group['Source'].iloc[0] == 'Kraken' and any('Order:' in str(note) for note in group['Notes']):
        # Extract ordertxid (should be the same for all rows in group)
        ordertxid = next(note.split('Order:')[1].split()[0] 
                        for note in group['Notes'] if 'Order:' in str(note))
        
        # Convert IDs to strings before joining
        trade_ids = ' | '.join(str(id_) for id_ in group['ID'])
        
        # Combine notes in order: Order, Trades, Group info
        combined_notes = f"From Order: {ordertxid} ~~ Grouped {len(group)} Trades: {trade_ids} ~~ Tx IDs: {' | '.join(str(id_) for id_ in group['ID'])}"
    else:
        # For non-Kraken transactions, preserve original notes
        original_notes = [note for note in group['Notes'] if note]
        grouped_note = f"Grouped {len(group)} Tx IDs: {' | '.join(str(id_) for id_ in group['ID'])}"
        combined_notes = ' ~~ '.join(filter(None, original_notes + [grouped_note]))
    
    # Create combined transaction
    combined = {
        'ID': f"{str(group['ID'].iloc[0])}_grouped",
        'Timestamp': group['Timestamp'].iloc[0],
        'Source': group['Source'].iloc[0],
        'Type': group['Type'].iloc[0],
        'Asset': group['Asset'].iloc[0],
        'Amount': float(amount_sum),
        'Subtotal': float(subtotal_sum),
        'Fee': float(fee_sum),
        'Total USD': float(total_usd_sum),
        'Spot Price': float(weighted_spot_price),
        'Notes': combined_notes
    }
    
    return pd.Series(combined)

def get_historical_price(asset: str, timestamp: datetime) -> float:
    # Look up the historical price of a cryptocurrency at a specific date/time
    # First checks cache, then uses CryptoCompare API if needed
    global price_cache
    
    # Format date for display
    date_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
    
    # Handle LUNA/LUNA2 transition
    display_name = asset
    if asset == 'LUNA':
        # After transition date, original LUNA became LUNC
        if timestamp >= LUNA_TRANSITION_DATE:
            asset = 'LUNC'
            display_name = 'LUNA (AKA: LUNC)'
    elif asset == 'LUNA2':
        # After transition date, LUNA2 price data is under LUNA
        if timestamp >= LUNA_TRANSITION_DATE:
            asset = 'LUNA'
            display_name = 'LUNA2 (AKA: LUNA)'
    
    print(f"\nLooking up price for {display_name} at {date_str}")
    
    # Determine if we can use minute-level data
    use_minute_data = CRYPTOCOMPARE_API_KEY and CRYPTOCOMPARE_API_KEY != 'YOUR-API-KEY-HERE'
    
    # Round timestamp based on data granularity
    if use_minute_data:
        rounded_timestamp = timestamp.replace(second=0, microsecond=0)
        cache_key = f"{asset}_{rounded_timestamp.strftime('%Y-%m-%d_%H-%M')}"
    else:
        rounded_timestamp = timestamp.replace(hour=0, minute=0, second=0, microsecond=0)
        cache_key = f"{asset}_{rounded_timestamp.strftime('%Y-%m-%d')}"
    
    no_price_key = f"no_price_{asset}_{rounded_timestamp.strftime('%Y-%m-%d')}"
    
    # Check regular price cache first
    if cache_key in price_cache:
        cached_data = price_cache[cache_key]
        if isinstance(cached_data, dict) and 'price' in cached_data:
            # If cache entry has timestamp, check if it's recent
            if 'timestamp' in cached_data:
                cache_age = datetime.now() - datetime.fromtimestamp(cached_data['timestamp'])
                if cache_age.days < 30:  # Cache valid for 30 days
                    print(f"Found cached price for {display_name}: ${cached_data['price']:.8f}")
                    return float(cached_data['price'])
            else:
                # Legacy cache format - just return the price
                print(f"Found legacy cached price for {display_name}: ${float(cached_data['price']):.8f}")
                return float(cached_data['price'])
        elif isinstance(cached_data, (int, float)):
            # Direct price storage format
            print(f"Found simple cached price for {display_name}: ${float(cached_data):.8f}")
            return float(cached_data)
    
    # Check no-price cache
    if no_price_key in price_cache:
        cached_data = price_cache[no_price_key]
        if isinstance(cached_data, dict) and 'timestamp' in cached_data:
            cache_age = datetime.now() - datetime.fromtimestamp(cached_data['timestamp'])
            if cache_age.total_seconds() < 43200:  # No-price cache valid for 12 hours
                print(f"No price available for {display_name} (cached response)")
                return 0
        # If cache format is invalid or expired, remove it
        del price_cache[no_price_key]
    
    print(f"Fetching price from API for {display_name} at {date_str}...")
    
    try:
        # Convert timestamp to UNIX timestamp
        unix_time = int(rounded_timestamp.timestamp())
        
        # Try minute data first if available
        if use_minute_data and (datetime.now() - rounded_timestamp).days <= 7:
            url = f"https://min-api.cryptocompare.com/data/v2/histominute"
            params = {
                'fsym': asset,
                'tsym': 'USD',
                'limit': 1,
                'toTs': unix_time
            }
            headers = {'authorization': CRYPTOCOMPARE_API_KEY}
        else:
            # Fall back to daily data
            url = f"https://min-api.cryptocompare.com/data/v2/histoday"
            params = {
                'fsym': asset,
                'tsym': 'USD',
                'limit': 1,
                'toTs': unix_time
            }
            headers = {}
        
        # Add delay to avoid rate limiting
        time.sleep(0.25)
        
        # Make API request
        response = requests.get(url, params=params, headers=headers)
        data = response.json()
        
        if response.status_code != 200:
            log_error(f"API Error ({response.status_code}): {data.get('Message', 'Unknown error')}")
            return 0
        
        # Try to get price from response
        if data.get('Response') == 'Success' and data.get('Data', {}).get('Data'):
            close_price = data['Data']['Data'][0].get('close', 0)
            
            # If no price found, try previous day
            if close_price == 0 and not use_minute_data:
                yesterday = rounded_timestamp - timedelta(days=1)
                unix_time = int(yesterday.timestamp())
                params['toTs'] = unix_time
                
                time.sleep(0.25)  # Add delay for second request
                response = requests.get(url, params=params, headers=headers)
                data = response.json()
                
                if response.status_code == 200 and data.get('Response') == 'Success' and data.get('Data', {}).get('Data'):
                    close_price = data['Data']['Data'][0].get('close', 0)
            
            # Cache result if we found a price
            if close_price > 0:
                print(f"Found price for {display_name}: ${close_price:.8f}")
                price_cache[cache_key] = {
                    'price': float(close_price),
                    'timestamp': int(datetime.now().timestamp())
                }
                save_price_cache(price_cache)
                return float(close_price)
        
        # If we get here, no price was found
        msg = f"No price data found for {display_name} at {rounded_timestamp}"
        log_error(msg)
        print(msg)
        
        # Cache the no-price result
        price_cache[no_price_key] = {
            'timestamp': int(datetime.now().timestamp())
        }
        save_price_cache(price_cache)
        return 0
        
    except requests.exceptions.RequestException as e:
        msg = f"Network error fetching price for {display_name}: {str(e)}"
        log_error(msg)
        print(msg)
        return 0
        
    except Exception as e:
        msg = f"Unexpected error getting price for {display_name}: {str(e)}"
        log_error(msg)
        print(msg)
        return 0

def strip_trailing_zeros(num) -> str:
    # Convert number to string, removing trailing zeros after decimal while preserving significant digits
    try:
        # Convert input to float first if it's a string
        if isinstance(num, str):
            num = float(num)

        # Convert to string with high precision to preserve all significant digits
        str_val = f"{float(num):.10f}"

        # Remove trailing zeros after decimal point, but keep decimal if whole number
        return str_val.rstrip('0').rstrip('.') + ('0' if str_val.endswith('.0') else '')

    except Exception as e:
        print(f"Error in strip_trailing_zeros for value '{num}': {str(e)}")
        return str(num)  # Return original value as string if conversion fails

def update_missing_spot_prices(df: pd.DataFrame) -> pd.DataFrame:
    # Find any transactions missing price data and look up their prices
    # Also calculates missing subtotals using the newly found prices
    df = df.copy()
    historical_ids = set()  # Track transaction IDs instead of indices
    calculated = {
        'Subtotal': set(),  # Track subtotals calculated from historical prices
        'Fee': set(),
        'Total USD': set(),
        'Spot Price': set()
    }
    
    # Find rows with missing spot prices
    missing_prices = (df['Spot Price'] == 0) & (df['Asset'] != 'USD')
    missing_count = missing_prices.sum()
    
    if missing_count > 0:
        print(f"\nFetching {missing_count} missing spot prices...")
        
        # Track progress
        current = 0
        for idx, row in df[missing_prices].iterrows():
            current += 1
            
            if row['Asset'] != 'USD':
                # Store original print function
                old_print = print
                printed_messages = []
                
                # Create temporary print function to capture output
                def temp_print(*args, **kwargs):
                    message = ' '.join(str(arg) for arg in args)
                    printed_messages.append(message)
                
                # Replace print function
                globals()['print'] = temp_print
                
                # Get the price
                price = get_historical_price(row['Asset'], row['Timestamp'])
                
                # Restore print function
                globals()['print'] = old_print
                
                # Determine status based on captured messages
                if price > 0:
                    if any('Found cached price' in msg or 'Found legacy cached price' in msg or 'Found simple cached price' in msg for msg in printed_messages):
                        status = "CACHED SUCCESS"
                    elif any('Found price for' in msg for msg in printed_messages):
                        status = "FETCH SUCCESS"
                    else:
                        status = "SUCCESS"
                else:
                    status = "FAILED"
                
                # Format price string with significant digits
                if price == 0:
                    price_str = "$0.00"
                else:
                    # Strip trailing zeros but preserve significant digits
                    price_str = f"${strip_trailing_zeros(price)}"
                
                # Print single summary line
                print(f"Processing {current}/{missing_count}: {row['Asset']} at {row['Timestamp']} - {status} - Price: {price_str}")
                
                if price > 0:
                    # Store the price with full precision but stripped of trailing zeros
                    df.loc[idx, 'Spot Price'] = float(strip_trailing_zeros(price))
                    historical_ids.add(row['ID'])  # Track transaction ID instead of index
    
    # After updating spot prices, recalculate missing subtotals
    missing_subtotals = (
        (df['Subtotal'] == 0) & 
        (df['Amount'].fillna(0) != 0) & 
        (df['Spot Price'].fillna(0) != 0)
    )
    
    if missing_subtotals.any():
        print(f"\nRecalculating {missing_subtotals.sum()} missing subtotals...")
        df.loc[missing_subtotals, 'Subtotal'] = (
            df.loc[missing_subtotals, 'Amount'] * 
            df.loc[missing_subtotals, 'Spot Price']
        )
        
        # Track transaction IDs for subtotals calculated from historical prices
        for idx in df[missing_subtotals].index:
            if df.loc[idx, 'ID'] in historical_ids:
                calculated['Subtotal'].add(df.loc[idx, 'ID'])
                
                # If Fee is 0, copy Subtotal to Total USD
                row = df.loc[idx]
                if (row['Fee'] == 0):
                    df.loc[idx, 'Total USD'] = df.loc[idx, 'Subtotal']
                    calculated['Total USD'].add(row['ID'])
    
    return df, historical_ids, calculated

def merge_all_transactions():
    # Main function that combines all cryptocurrency transactions:
    # 1. Loads transactions from all exchange files
    # 2. Standardizes the data format and fixes any missing information
    # 3. Groups related transactions that happen close together in time
    # 4. Saves everything to ALL-MASTER-crypto-transactions.xlsx
    try:
        print("\n\nStarting merge process...\n")
        
        # Load all transaction data
        coinbase_df = load_coinbase_transactions()
        print(f"")
        coinbase_pro_df = load_coinbase_pro_transactions()
        print(f"")
        kraken_df = load_kraken_transactions()
        print(f"")
        strike_df = load_strike_transactions()
        print(f"")
        cashapp_df = load_cashapp_transactions()
        print(f"")
        
        # Check for existing manual transactions
        manual_file = Path('add-manual-transactions.xlsx')
        manual_df = pd.DataFrame()
        if manual_file.exists():
            try:
                manual_df = read_transaction_file(manual_file)
                if not manual_df.empty:
                    print("\n **  Loading Manual transactions...")
                    manual_df['Timestamp'] = pd.to_datetime(manual_df['Timestamp'])
                    manual_df['Source'] = manual_df['Source'].fillna('Manual')
                    print(f"\nFound {len(manual_df)} manual transactions to process")
            except Exception as e:
                print(f"Error reading manual transactions: {str(e)}")
                manual_df = pd.DataFrame()
        
        # Just do basic NaN filling for now - no calculations yet
        all_dfs = []
        for df in [coinbase_df, coinbase_pro_df, kraken_df, strike_df, cashapp_df, manual_df]:
            if not df.empty:
                df = df.copy()
                df['Notes'] = df['Notes'].fillna('')
                df['Fee'] = df['Fee'].fillna(0)
                df['Spot Price'] = df['Spot Price'].fillna(0)
                df['Subtotal'] = df['Subtotal'].fillna(0)
                df['Total USD'] = df['Total USD'].fillna(0)
                all_dfs.append(df)
        
        if all(df.empty for df in all_dfs):
            print("Error: No transaction data found in any files.")
            return
        
        # Force timezone-naive for all DataFrames before concat
        for df in all_dfs:
            if not df.empty and 'Timestamp' in df.columns:
                df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.tz_localize(None)
        
        # Define the standard columns used throughout the script
        # This is used by the manual transaction template and final output
        required_columns = [
            'ID', 'Timestamp', 'Source', 'Type', 'Asset', 'Amount',
            'Subtotal', 'Fee', 'Total USD', 'Spot Price', 'Notes'
        ]
        
        # Combine all transactions
        all_transactions = safe_concat(all_dfs, columns=required_columns, ignore_index=True)
        
        # Standardize signs
        all_transactions = standardize_transaction_values(all_transactions)
        
        # Sort by timestamp
        all_transactions = all_transactions.sort_values('Timestamp')
        
        # Clean up and standardize
        all_transactions['Timestamp'] = pd.to_datetime(all_transactions['Timestamp'])
        
        # Convert numeric columns
        numeric_columns = ['Amount', 'Subtotal', 'Total USD', 'Fee']
        for col in numeric_columns:
            all_transactions[col] = pd.to_numeric(all_transactions[col], errors='coerce')
        
        # Clean and convert spot prices to numeric values
        all_transactions['Spot Price'] = all_transactions['Spot Price'].apply(clean_price_string)
        
        # Group transactions first
        all_transactions = assign_lot_ids_and_group(all_transactions)
        
        # NOW calculate missing values and track what was calculated
        all_transactions, calculated = fill_missing_transaction_values(all_transactions)
        calculated_cells.update(calculated)  # Update our global tracking dict
        
        # Update missing spot prices and track historical prices
        all_transactions, historical_cells, hist_calculated = update_missing_spot_prices(all_transactions)
        
        # Clean up and standardize
        all_transactions['Timestamp'] = pd.to_datetime(all_transactions['Timestamp'])
        
        # Convert numeric columns
        numeric_columns = ['Amount', 'Subtotal', 'Total USD', 'Fee']
        for col in numeric_columns:
            all_transactions[col] = pd.to_numeric(all_transactions[col], errors='coerce')
        
        # Clean spot prices
        all_transactions['Spot Price'] = all_transactions['Spot Price'].apply(clean_price_string)
        
        # Merge the calculated cells from historical prices
        for col, indices in hist_calculated.items():
            calculated_cells[col].update(indices)
     
        # Reorder columns
        # Add Lot ID to required columns for final output
        final_columns = required_columns[:-1] + ['Lot ID'] + [required_columns[-1]]
        
        # Now reorder after all columns exist
        all_transactions = all_transactions[final_columns]
        
        print(f"\nWriting to Excel file...")
        
        # Ensure USD values are negative for buys and positive for sells
        for idx, row in all_transactions.iterrows():
            if 'buy' in row['Type'].lower():
                all_transactions.at[idx, 'Total USD'] = -abs(all_transactions.at[idx, 'Total USD'])
                all_transactions.at[idx, 'Subtotal'] = -abs(all_transactions.at[idx, 'Subtotal'])
            elif 'sell' in row['Type'].lower():
                all_transactions.at[idx, 'Total USD'] = abs(all_transactions.at[idx, 'Total USD'])
                all_transactions.at[idx, 'Subtotal'] = abs(all_transactions.at[idx, 'Subtotal'])

        # Save to Excel
        with pd.ExcelWriter('ALL-MASTER-crypto-transactions.xlsx', engine='openpyxl') as writer:
            # Convert only Amount column to string to preserve full precision
            numeric_df = all_transactions.copy()
            
            # Only convert Amount to string with full precision
            if 'Amount' in numeric_df.columns:
                numeric_df['Amount'] = numeric_df['Amount'].apply(
                    lambda x: format(Decimal(str(x)), 'f') if pd.notnull(x) else ''
                )
            
            numeric_df.to_excel(writer, index=False, sheet_name='Transactions')
            
            # Get the worksheet
            worksheet = writer.sheets['Transactions']

            # Format worksheet
            format_excel_worksheet(worksheet, df=all_transactions, sheet_name='Transactions', 
                               calculated_cells=calculated_cells,
                               historical_cells=historical_cells)
            
        print(f"   DONE!!\n\nFile saved as...\n   ALL-MASTER-crypto-transactions.xlsx")
        
        # Create manual transactions template if it doesn't exist
        if not manual_file.exists():
            print("   ***********************************")
            print("\nCreating manual transactions template file...")
            template_df = pd.DataFrame(columns=[
                'ID', 'Timestamp', 'Source', 'Type', 'Asset', 'Amount',
                'Subtotal', 'Fee', 'Total USD', 'Spot Price', 'Notes'
            ])
            
            with pd.ExcelWriter(manual_file, engine='openpyxl') as writer:
                template_df.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']
                
                # Add filters
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Set column widths
                column_widths = {
                    'ID': 25,
                    'Timestamp': 20,
                    'Source': 15,
                    'Type': 15,
                    'Asset': 10,
                    'Amount': 15,
                    'Subtotal': 15,
                    'Fee': 10,
                    'Total USD': 15,
                    'Spot Price': 15,
                    'Notes': 50
                }
                
                for i, col in enumerate(template_df.columns):
                    worksheet.column_dimensions[get_column_letter(i + 1)].width = column_widths.get(col, 15)
                
                # Format header row
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
            
            print("\n ... Created 'add-manual-transactions.xlsx'")
            print("              ****************************")
            print("\n  ** NOTICE **: ADD any transactions from 'unsupported' exchanges to this 'manual' file.")
            print("\n######################################################################################")
            print("\nRequired fields:")
            print("  - Timestamp: When the transaction occurred")
            print("  - Type: Buy, Sell, Send, Receive, etc.")
            print("  - Asset: The cryptocurrency symbol (BTC, ETH, etc.)")
            print("  - Amount: How much crypto was traded")
            print("\nHelpful fields (the more you provide, the better):")
            print("  - Subtotal: Total USD before fees")
            print("  - Fee: Transaction fee amount")
            print("  - Total USD: Total cost/proceeds including fees")
            print("  - Spot Price: Price per coin at time of transaction")
            print("\nExtra fields (for your own reference):")
            print("  - ID: The unique identifier for the transaction")
            print("  - Source: The source of the transaction (Binance, Bitfinex, Gemini, etc.)")
            print("  - Notes: Any additional information about the transaction")
            print("\nThe script will calculate missing values more accurately if you provide as much information as possible.")
            print("\n##################################################################################")
            print("\n  ##########")
            print("  * NOTICE *:  If you have *ADDITIONAL* transactions from...")
            print("  ##########      - other exchanges...")
            print("                  - private sales...")
            print("                  - Bitcoin ATMs... etc.")
            print("\n               Then ADD these extra transactions to the newly created 'add-manual-transactions.xlsx' file")
            print("                  ... and RUN this '1-merge_crypto_txs.py' script AGAIN.")
            print("\n########################################################################################")
            print("\n   If you do *NOT* have additional transactions to manually add...")
            print("\n   - You can view the 'ALL-MASTER-crypto-transactions.xlsx' file in Excel, LibreOffice, or Google Sheets.")
            print("                       ***********************************")
            print("   - You can also run the '2-calculate-tax-lots.py' script to calculate your tax lots.")
            print("                           ***********************")
            print("\n  **IMPORTANT**:  Make sure to BACK UP your 'ALL-MASTER-crypto-transactions.xlsx' file!")
            print("                               **** **       ***********************************")
            print("###########################################################################################")

    except Exception as e:
        print(f"Error merging transactions: {str(e)}")


def format_excel_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame = None, sheet_name: str = '', calculated_cells=None, historical_cells=None):
    # Make the Excel file easy to read and use:
    # - Blue headers with filters for sorting
    # - Proper column widths based on content
    # - Applies appropriate number formats (currency, dates, decimals)
    # - Adds filters to the worksheet for easy sorting and filtering
    # - Colors negative numbers in red

    # Add filters
    if df is not None:
        worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df) + 1}"
    
    # Define styles
    header_fill = openpyxl.styles.PatternFill(
        start_color='CCE5FF',  # Light blue
        end_color='CCE5FF',
        fill_type='solid'
    )
    header_font = openpyxl.styles.Font(bold=True)
    center_align = openpyxl.styles.Alignment(horizontal='center')
    left_align = openpyxl.styles.Alignment(horizontal='left')
    red_font = openpyxl.styles.Font(color='FF0000')  # Red color for negative numbers
    
    # First pass: Apply basic header formatting and calculate content lengths
    column_content_lengths = {}
    
    for column in worksheet.columns:
        col_letter = column[0].column_letter
        header_cell = column[0]
        header_value = str(header_cell.value)
        
        # Apply basic header formatting
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = center_align  # Default to center
        
        # Calculate max content length (excluding header)
        max_content_length = 0
        for cell in column[1:]:  # Skip header
            try:
                if cell.value:
                    max_content_length = max(max_content_length, len(str(cell.value)))
            except:
                pass
        column_content_lengths[col_letter] = max_content_length
    
    # Second pass: Handle column widths, alignments, and number formats
    for column in worksheet.columns:
        col_letter = column[0].column_letter
        header_cell = column[0]
        header_value = str(header_cell.value)
        header_length = len(header_value)
        max_content_length = column_content_lengths[col_letter]
        
        # Calculate base widths
        content_width = max_content_length + 2  # Add minimal padding
        header_width = header_length + 4  # Add space for filter dropdown
        
        # Use the larger of content or header width, with minimum of 10
        base_width = max(content_width, header_width, 10)
        
        # Reduce width for monetary columns
        if header_value in ['Subtotal', 'Fee', 'Total USD', 'Spot Price']:
            base_width = max(base_width - 6, 10)  # Subtract 5, but keep minimum of 10
        
        # Special handling for Notes column
        if header_value == 'Notes':
            final_width = min(base_width, 200)  # Cap Notes column at 200
        else:
            final_width = min(base_width, 100)  # Other columns capped at 100
        
        # Set column width
        worksheet.column_dimensions[col_letter] = openpyxl.worksheet.dimensions.ColumnDimension(
            worksheet, 
            index=col_letter, 
            width=final_width,
            bestFit=False
        )
        
        # Apply number formatting to data cells
        header_value_lower = header_value.lower()
        for cell in column[1:]:  # Skip header
            if any(x in header_value_lower for x in ['price', 'usd', 'subtotal', 'fee']):
                cell.number_format = '_($* #,##0.00_);[Red]_($* -#,##0.00_)'  # Left-aligned $ with red minus
            elif header_value == 'Amount':  # Exact match for Amount column
                if cell.value is not None:
                    # Count significant decimal places in the actual value
                    str_val = str(cell.value)
                    if '.' in str_val:
                        decimal_part = str_val.split('.')[1]
                        # Find the last non-zero digit position
                        last_nonzero = len(decimal_part)
                        for i in range(len(decimal_part) - 1, -1, -1):
                            if decimal_part[i] != '0':
                                last_nonzero = i + 1
                                break
                        decimals = last_nonzero
                        if decimals > 0:
                            cell.number_format = f'#,##0.{"0" * decimals};[Red]-#,##0.{"0" * decimals}'
                        else:
                            cell.number_format = '#,##0;[Red]-#,##0'
                    else:
                        cell.number_format = '#,##0;[Red]-#,##0'
            elif 'date' in header_value_lower:
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            elif 'timestamp' in header_value_lower:
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                cell.alignment = openpyxl.styles.Alignment(horizontal='left')

    # Add highlighting for calculated and historical values
    if calculated_cells and df is not None:
        for col, tx_ids in calculated_cells.items():
            if tx_ids:
                col_idx = df.columns.get_loc(col) + 1
                for row_idx, row in df.iterrows():
                    if row['ID'] in tx_ids:
                        cell = worksheet.cell(row=row_idx + 2, column=col_idx)
                        cell.fill = CALCULATED_FILL
    
    if historical_cells and df is not None:
        price_col_idx = df.columns.get_loc('Spot Price') + 1
        for row_idx, row in df.iterrows():
            if row['ID'] in historical_cells:
                cell = worksheet.cell(row=row_idx + 2, column=price_col_idx)
                cell.fill = HISTORICAL_FILL

def clear_old_cache_entries(days_old: int = 30):
    # Clean up the price cache by removing old entries
    # By default, removes entries older than 30 days
    cache = load_price_cache()
    current_time = datetime.now()
    
    # Filter out old entries
    new_cache = {}
    for key, value in cache.items():
        try:
            # Extract date from cache key
            date_str = key.split('_')[1]
            entry_date = datetime.strptime(date_str, '%Y-%m-%d')
            
            # Keep if within time window
            if (current_time - entry_date).days <= days_old:
                new_cache[key] = value
        except:
            continue
    
    save_price_cache(new_cache)

def should_merge_transactions(tx1: Dict, tx2: Dict, time_threshold: int = 90) -> bool:
    # Determine if two transactions are part of the same trade:
    # - Must be same asset and type (buy/sell)
    # - Must happen within 90 seconds of each other
    # - For Kraken, if trades file present: must share the same order ID
    # Don't merge Receive or Learning Reward transactions
    if tx1['Type'] in ['Receive', 'Learning Reward'] or tx2['Type'] in ['Receive', 'Learning Reward']:
        return False
    
    # Don't merge if either transaction is a Convert (they're already paired)
    if 'Converted' in tx1.get('Notes', '') or 'Converted' in tx2.get('Notes', ''):
        return False
    
    # Check if these are Kraken transactions
    if tx1['Source'] == 'Kraken' and tx2['Source'] == 'Kraken':
        # Only proceed with ordertxid matching if both have Notes with Order
        if ('Notes' in tx1 and 'Notes' in tx2 and 
            'Order:' in tx1['Notes'] and 'Order:' in tx2['Notes']):
            
            # Extract ordertxid from Notes
            ordertxid1 = tx1['Notes'].split('Order:')[1].split()[0]
            ordertxid2 = tx2['Notes'].split('Order:')[1].split()[0]
            
            # For Kraken, only check ordertxid match and basic criteria (ignore time)
            return (
                ordertxid1 == ordertxid2 and
                tx1['Type'] == tx2['Type'] and
                tx1['Asset'] == tx2['Asset'] and
                ((tx1['Amount'] > 0) == (tx2['Amount'] > 0))  # Same sign
            )
    
    # For non-Kraken transactions, use time-based merging
    time_diff = abs((tx2['Timestamp'] - tx1['Timestamp']).total_seconds())
    
    return (
        time_diff <= time_threshold and
        tx1['Type'] == tx2['Type'] and
        tx1['Asset'] == tx2['Asset'] and
        tx1['Source'] == tx2['Source'] and
        ((tx1['Amount'] > 0) == (tx2['Amount'] > 0))  # Same sign
    )

def fill_missing_transaction_values(df: pd.DataFrame) -> pd.DataFrame:
    # Fill in missing transaction values and standardize data:
    # 1. Replace NaN/zero values with calculated values where possible
    # 2. Ensure consistent signs for monetary values
    # 3. Handle all combinations of available data
    if df.empty:
        return df, {}
    
    df = df.copy()
    calculated = {
        'Subtotal': set(),
        'Fee': set(),
        'Total USD': set(),
        'Spot Price': set()
    }
    
    # First pass:  Fill NaN values with appropriate defaults
    df['Notes'] = df['Notes'].fillna('')
    df['Fee'] = df['Fee'].fillna(0)
    df['Spot Price'] = df['Spot Price'].fillna(0)
    df['Subtotal'] = df['Subtotal'].fillna(0)
    df['Total USD'] = df['Total USD'].fillna(0)
    
    # Second pass:  Calculate Spot Price where missing
    spot_price_mask = (df['Spot Price'] == 0) & (df['Amount'].fillna(0) != 0)
    
    # Try Subtotal first, then (Total USD - Fee), and Total USD as last resort
    subtotal_mask = spot_price_mask & (df['Subtotal'].fillna(0) != 0)
    total_fee_mask = spot_price_mask & ~subtotal_mask & (df['Total USD'].fillna(0) != 0) & (df['Fee'].fillna(0) != 0)
    total_usd_mask = spot_price_mask & ~subtotal_mask & ~total_fee_mask & (df['Total USD'].fillna(0) != 0)
    
    if subtotal_mask.any():
        df.loc[subtotal_mask, 'Spot Price'] = df.loc[subtotal_mask, 'Subtotal'].abs() / df.loc[subtotal_mask, 'Amount'].abs()
        calculated['Spot Price'].update(df.loc[subtotal_mask, 'ID'])
    
    if total_fee_mask.any():
        df.loc[total_fee_mask, 'Spot Price'] = (df.loc[total_fee_mask, 'Total USD'] - df.loc[total_fee_mask, 'Fee']).abs() / df.loc[total_fee_mask, 'Amount'].abs()
        calculated['Spot Price'].update(df.loc[total_fee_mask, 'ID'])
    
    if total_usd_mask.any():
        df.loc[total_usd_mask, 'Spot Price'] = df.loc[total_usd_mask, 'Total USD'].abs() / df.loc[total_usd_mask, 'Amount'].abs()
        calculated['Spot Price'].update(df.loc[total_usd_mask, 'ID'])
    
    # Third pass:  Calculate Fee where missing but we have Total USD and Subtotal
    fee_mask = (df['Fee'] == 0) & (df['Total USD'] != 0) & (df['Subtotal'] != 0)
    if fee_mask.any():
        calculated_fees = df.loc[fee_mask, 'Total USD'].abs() - df.loc[fee_mask, 'Subtotal'].abs()
        # Only update fees that are actually non-zero
        non_zero_fees = calculated_fees != 0
        if non_zero_fees.any():
            df.loc[fee_mask[fee_mask].index[non_zero_fees], 'Fee'] = calculated_fees[non_zero_fees]
            calculated['Fee'].update(df.loc[fee_mask[fee_mask].index[non_zero_fees], 'ID'])
    
    # Calculate Subtotal where missing but we have enough info
    subtotal_mask = (df['Subtotal'] == 0)
    # From Total USD and Fee
    total_fee_mask = subtotal_mask & (df['Total USD'] != 0) & (df['Fee'] != 0)
    if total_fee_mask.any():
        df.loc[total_fee_mask, 'Subtotal'] = df.loc[total_fee_mask, 'Total USD'] - df.loc[total_fee_mask, 'Fee'].abs()
        calculated['Subtotal'].update(df.loc[total_fee_mask, 'ID'])
    
    amount_price_mask = subtotal_mask & ~total_fee_mask & (df['Amount'] != 0) & (df['Spot Price'] != 0)
    if amount_price_mask.any():
        df.loc[amount_price_mask, 'Subtotal'] = df.loc[amount_price_mask, 'Amount'] * df.loc[amount_price_mask, 'Spot Price']
        calculated['Subtotal'].update(df.loc[amount_price_mask, 'ID'])
        
        # For rows where we calculated Subtotal from historical price and Fee is 0,
        # copy Subtotal to Total USD
        zero_fee_mask = amount_price_mask & (df['Fee'] == 0)
        if zero_fee_mask.any():
            df.loc[zero_fee_mask, 'Total USD'] = df.loc[zero_fee_mask, 'Subtotal']
            calculated['Total USD'].update(df.loc[zero_fee_mask, 'ID'])
    
    # Finally, calculate any remaining missing Total USD from Subtotal and Fee
    total_mask = (df['Total USD'] == 0)
    if total_mask.any():
        df.loc[total_mask, 'Total USD'] = df.loc[total_mask, 'Subtotal'].abs() + df.loc[total_mask, 'Fee'].abs()
        calculated['Total USD'].update(df.loc[total_mask, 'ID'])
    
    return df, calculated

def main():
    merge_all_transactions()

def read_transaction_file(file_path):
    #Attempts to read a transaction file in either XLSX or CSV format.
    #Returns a pandas DataFrame or empty DataFrame if file can't be read.
    try:
        # Convert Path object to string for extension checking
        file_path_str = str(file_path)
        
        # Try Excel first
        if file_path_str.lower().endswith('.xlsx'):
            return pd.read_excel(file_path)
        # Try CSV with different encodings and delimiters
        elif file_path_str.lower().endswith('.csv'):
            # For Coinbase CSV, we need to find the actual header row
            if 'coinbase' in file_path_str.lower() and 'pro' not in file_path_str.lower():
                try:
                    # Read first few rows to find headers
                    temp_df = pd.read_csv(file_path, nrows=20)
                    for i in range(len(temp_df)):
                        row_values = temp_df.iloc[i].values
                        # Look for multiple header indicators
                        if any('Timestamp' in str(x) for x in row_values) or \
                           any('Transaction Type' in str(x) for x in row_values) or \
                           any('Quantity Transacted' in str(x) for x in row_values):
                            # Get the actual column names from this row
                            column_names = [str(x).strip() for x in row_values]
                            # Read the file again, skipping to the data and using our column names
                            return pd.read_csv(file_path, 
                                          skiprows=i+1,    # Skip past the header row
                                          names=column_names)  # Use the column names we found
                except:
                    pass
            
            # For other CSV files, try different encodings
            try:
                return pd.read_csv(file_path, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    return pd.read_csv(file_path, encoding='utf-8-sig')
                except:
                    try:
                        return pd.read_csv(file_path, encoding='latin1')
                    except:
                        print(f"\nDEBUG: Failed to read {file_path} with any encoding")
                        return pd.DataFrame()
        return pd.DataFrame()
    except Exception as e:
        print(f"Error reading file {file_path}: {str(e)}")
        return pd.DataFrame()

if __name__ == "__main__":
    main()
