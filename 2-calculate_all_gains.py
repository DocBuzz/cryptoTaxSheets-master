# This script calculates cryptocurrency capital gains/losses using various accounting methods
# It processes transaction history and determines the cost basis and gains for each sale

# API key for CryptoCompare - used to fetch historical cryptocurrency prices
# If no API key is provided, the script will use daily historical prices
# With an API key, we can get minute-by-minute prices for the last 7 days
CRYPTOCOMPARE_API_KEY = 'YOUR-API-KEY-HERE'

# Import required libraries for data processing, date handling, and calculations

import pandas as pd                             # For data manipulation and analysis
from datetime import datetime, timedelta        # For handling dates and times
from decimal import Decimal, ROUND_HALF_UP      # For precise decimal calculations
from typing import List, Dict, Optional, Union  # For type hints to make code more reliable
import logging                                  # For error tracking and debugging
from dataclasses import dataclass               # For creating structured data classes
from enum import Enum                           # For creating enumerated types
import re                                       # For text pattern matching
from collections import defaultdict             # For creating dictionaries with default values
import requests                                 # For making HTTP requests to get price data
import time                                     # For adding delays between API calls
import json                                     # For reading/writing cache files
from pathlib import Path                        # For handling file paths
import openpyxl                                 # For Excel file operations
from openpyxl.styles import PatternFill         # For adding conditional formatting

# Important date when LUNA became LUNC (Luna Classic)
LUNA_TRANSITION_DATE = datetime(2022, 5, 28)

# Important date constant: When NU stopped working on Coinbase (Feb 6, 2023). 1 NU is worth 3.26 T
NU_TRANSITION_DATE = datetime(2023, 2, 6)

# Set up logging to only create the log file if an error is actually logged
class ErrorOnlyFileHandler(logging.FileHandler):
    #Only creates the log file if an error is actually logged
    def __init__(self, filename, mode='a', encoding=None, delay=True):  # Set delay=True
        super().__init__(filename, mode, encoding, delay)
        self.error_occurred = False

    def emit(self, record):
        if record.levelno >= logging.ERROR:  # Only handle ERROR or higher
            self.error_occurred = True
            super().emit(record)

handler = ErrorOnlyFileHandler('ALL-crypto-profit-and-loss-errors.log')
handler.setLevel(logging.ERROR)
handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

logger = logging.getLogger()
logger.setLevel(logging.ERROR)
logger.addHandler(handler)

def load_price_cache():
    # Load previously fetched cryptocurrency prices from the "price_cache.json" cache file.
    # Returns an empty dictionary if no cache exists or if there's an error reading it.

    cache_file = Path('price_cache.json')
    if cache_file.exists():
        try:
            with open(cache_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading price cache: {e}")
    return {}

def save_price_cache(cache):
    # Save fetched cryptocurrency prices to the "price_cache.json" cache file.

    try:
        with open('price_cache.json', 'w') as f:
            json.dump(cache, f)
    except Exception as e:
        print(f"Error saving price cache: {e}")

# Initialize price cache
price_cache = load_price_cache()

class TransactionType(Enum):
    # Define all possible types of cryptocurrency transactions.

    BUY = 'buy'
    SELL = 'sell'
    ADVANCED_TRADE_BUY = 'advanced trade buy'
    ADVANCED_TRADE_SELL = 'advanced trade sell'
    CONVERT = 'convert'
    STAKE = 'staking income'
    LEARNING_REWARD = 'learning reward'
    RECEIVE = 'receive'
    SEND = 'send'
    DEPOSIT = 'deposit'
    ADMIN_DEBIT = 'admin debit'
    WITHDRAWAL = 'withdrawal'
    GIFT = 'gift'
    DIVIDEND = 'dividend'
    PRO_WITHDRAWAL = 'pro withdrawal'
    PRO_DEPOSIT = 'pro deposit'
    
    @classmethod
    def from_string(cls, value: str) -> Optional['TransactionType']:
        # Convert string to TransactionType, handling various capitalizations
        # Handles special cases like internal transfers and advanced trades
        # Returns None for internal Coinbase transfers

        clean_value = value.lower().strip()
        
        # Handle special cases - return None for internal Coinbase transfers
        if any(x in clean_value for x in ['exchange withdrawal', 'pro withdrawal', 'pro deposit']):
            return None
        
        # Handle special cases
        if 'advanced trade' in clean_value:
            clean_value = clean_value.replace('advanced trade ', '')
        if clean_value in ['learning reward', 'receive']:
            return cls.GIFT
        if 'deposit' in clean_value:  # This won't catch 'pro deposit'
            return cls.DEPOSIT
        if 'withdrawal' in clean_value:  # This won't catch 'pro withdrawal'
            return cls.WITHDRAWAL
        if 'dividend' in clean_value:
            return cls.DIVIDEND
            
        # Try to match the cleaned value
        for member in cls:
            if member.value == clean_value:
                return member
            
        # Additional fuzzy matching
        if 'buy' in clean_value:
            return cls.BUY
        if 'sell' in clean_value:
            return cls.SELL
        if 'convert' in clean_value:
            return cls.CONVERT
        if 'stake' in clean_value or 'staking' in clean_value:
            return cls.STAKE
        if 'send' in clean_value:
            return cls.SEND
        if 'admin' in clean_value and 'debit' in clean_value:
            return cls.ADMIN_DEBIT
            
        raise ValueError(f"Unknown transaction type: {value}")

@dataclass
class AssetLot:
    # Represents a single purchase or acquisition of cryptocurrency (called a "lot").
    # Think of this like buying shares of stock - each purchase is a separate lot.
    #
    # For example: If you buy 1 BTC today and 0.5 BTC tomorrow, those are two separate lots.
    # This helps track the cost basis and holding period for each purchase separately.
    #
    # Fields:
    # - timestamp: When this lot was acquired
    # - amount: How much cryptocurrency was acquired
    # - cost_basis: How much was paid in USD
    # - source: Where it came from (e.g., 'Coinbase', 'Kraken')
    # - transaction_type: What kind of transaction (buy, gift, staking reward, etc.)
    # - remaining: How much of this lot is still available (not sold)
    # - transaction_id: Unique ID from the exchange
    # - lot_id: Our internal tracking ID
    # - gift_market_value: For gifts, tracks the market value when received

    timestamp: datetime
    amount: Decimal
    cost_basis: Decimal
    source: str
    transaction_type: TransactionType
    remaining: Decimal
    transaction_id: str
    lot_id: str
    gift_market_value: Optional[Decimal] = None
    
    def __post_init__(self):
        # After creating a new lot, calculate the cost per unit.
        # For gifts, we use the market value as the cost basis.
        # For regular purchases, we divide total cost by amount.

        if self.transaction_type == TransactionType.GIFT and self.gift_market_value is not None:
            self.cost_per_unit = (
                self.gift_market_value / self.amount if self.amount != 0 
                else Decimal('0')
            )
        else:
            self.cost_per_unit = (
                self.cost_basis / self.amount if self.amount != 0 
                else Decimal('0')
            )
    
    def get_age_in_days(self, reference_date: Optional[datetime] = None) -> int:
        # Calculate age of lot in days
        ref_date = reference_date or datetime.now()
        return (ref_date - self.timestamp).days
    
    def is_long_term_at_date(self, sale_date: Optional[datetime] = None) -> bool:
        # Determine if lot qualifies for long-term status
        return self.get_age_in_days(sale_date) >= 365

@dataclass
class SaleLotDetail:
    # Records the details of how a specific lot was used in a sale.
    # When you sell cryptocurrency, you might use multiple purchase lots.
    # This class tracks exactly how each lot contributed to the sale.
    #
    # For example: If you sell 1.5 BTC, you might use:
    # - 1.0 BTC from a lot bought in January
    # - 0.5 BTC from a lot bought in March
    #
    # Fields:
    # - lot_id: Which purchase lot was used
    # - amount_sold: How much was sold from this lot
    # - cost_basis: Original purchase cost for the amount sold
    # - proceeds: How much was received from selling this portion
    # - gain_loss: Profit or loss (proceeds minus cost_basis)
    # - is_long_term: Whether this qualifies as long-term gain/loss (held > 1 year)
    # - purchase_date: When this lot was originally bought
    # - holding_period_days: How many days between purchase and sale
    # - cost_basis_per_unit: Original purchase price per unit

    lot_id: str
    amount_sold: Decimal
    cost_basis: Decimal
    proceeds: Decimal
    gain_loss: Decimal
    is_long_term: bool
    purchase_date: datetime
    holding_period_days: int
    cost_basis_per_unit: Decimal

@dataclass
class SaleTransaction:
    # Records a complete sale transaction, which might use multiple lots.
    # This combines all the individual lot details into one complete sale record.
    #
    # For example: A sale of 1.5 BTC would include:
    # - The total amount sold (1.5 BTC)
    # - Total proceeds received
    # - Total cost basis from all lots used
    # - Whether the entire sale qualifies as long-term
    # - Details about each lot used (stored in lot_details)
    #
    # Fields:
    # - timestamp: When the sale occurred
    # - amount: Total amount of cryptocurrency sold
    # - proceeds: Total USD received from the sale
    # - cost_basis: Total original purchase cost of all lots used
    # - source: Where the sale occurred (e.g., 'Coinbase')
    # - transaction_id: Unique ID from the exchange
    # - lot_details: List of SaleLotDetail records for each lot used
    # - is_long_term: True only if ALL lots used were held > 1 year
    # - lots_sold: Comma-separated list of lot IDs used in this sale

    timestamp: datetime
    amount: Decimal
    proceeds: Decimal
    cost_basis: Decimal
    source: str
    transaction_id: str
    lot_details: List[SaleLotDetail]
    is_long_term: bool
    lots_sold: str

@dataclass
class StakingRecord:
    # Records cryptocurrency received as staking rewards.
    # Staking is like earning interest by holding cryptocurrency.
    # The IRS generally treats staking rewards as income when received.

    timestamp: datetime
    amount: Decimal
    market_value: Decimal
    source: str
    transaction_id: str

@dataclass
class GiftRecord:
    # Records cryptocurrency received as gifts or rewards (like Coinbase Learn rewards).
    # The IRS treats gifts differently from purchases - the cost basis can vary:
    # - If recipient sells for less than giver's basis: Use the lower sale price
    # - If recipient sells for more than giver's basis: Use giver's original basis

    timestamp: datetime
    amount: Decimal
    market_value: Decimal
    source: str
    transaction_id: str
    gift_type: str

class AccountingMethod(Enum):
    # Different methods for choosing which lots to sell first.
    # This choice can significantly impact your tax liability:
    #
    # FIFO (First In, First Out):
    # - Sells oldest lots first
    # - Most common and generally accepted by IRS
    # - Often results in more long-term gains
    #
    # LIFO (Last In, First Out):
    # - Sells newest lots first
    # - Can minimize gains in rising markets
    # - May result in more short-term gains
    #
    # HIFO (Highest In, First Out):
    # - Sells highest-cost lots first
    # - Minimizes gains or maximizes losses
    # - Most tax-efficient in many cases
    #
    # LOFO (Lowest In, First Out):
    # - Sells lowest-cost lots first
    # - Maximizes gains
    # - Rarely used (usually not tax-efficient)

    FIFO = "First In, First Out"
    LIFO = "Last In, First Out"
    HIFO = "Highest In, First Out"
    LOFO = "Lowest In, First Out"

class AssetTracker:
    # Manages all transactions for a single cryptocurrency (like BTC or ETH).
    # This is the main class that:
    # 1. Keeps track of all purchases (lots)
    # 2. Records all sales and their details
    # 3. Tracks staking income and gifts
    # 4. Handles transfers between exchanges
    #
    # Think of this as your complete transaction history for one cryptocurrency.

    def __init__(self, asset_symbol: str, accounting_method: AccountingMethod):
        self.symbol = asset_symbol
        self.lots: List[AssetLot] = []                  # All purchase lots
        self.sales: List[SaleTransaction] = []          # All sales
        self.staking_income: List[StakingRecord] = []   # Staking rewards
        self.gifts: List[GiftRecord] = []               # Gifts and rewards
        self.sends: List[Dict] = []                     # Transfers sent to other wallets/exchanges
        self.deposits: List[Dict] = []                  # Transfers received
        self.transfers: List[Dict] = []                 # Internal transfers
        self.withdrawals: List[Dict] = []               # Withdrawals to external wallets
        self.accounting_method = accounting_method
        self.processed_sales = []                       # Track which sales we've handled
        self.errors = []                                # Track any issues that come up

    def add_lot(self, lot: AssetLot) -> None:
        # Add a new purchase lot to track.
        # If the lot doesn't have an ID, we generate one using:
        # - The cryptocurrency symbol (e.g., 'BTC')
        # - The timestamp of purchase
        # - A prefix 'GEN_' to show it's generated
        #
        # Example ID: GEN_BTC_20230615123456

        if pd.isna(lot.lot_id) or lot.lot_id == 'nan':
            # Generate a unique lot ID if none exists
            lot.lot_id = f"GEN_{self.symbol}_{lot.timestamp.strftime('%Y%m%d%H%M%S')}"
        self.lots.append(lot)

    def find_matching_send(self, deposit: Dict, window_hours: int = 8) -> Optional[Dict]:
        # Try to match a deposit with a previous send transaction.
        # This helps track transfers between exchanges.
        #
        # For example: If you send 1 BTC from Coinbase to Kraken:
        # 1. Look for a 'send' from Coinbase
        # 2. Look for a matching 'deposit' in Kraken
        # 3. Match them if they're within 8 hours and amounts match (allowing for fees)
        #
        # Parameters:
        # - deposit: The deposit transaction to match
        # - window_hours: How many hours to look back for matching send (default 8)

        deposit_time = deposit['timestamp']
        deposit_amount = deposit['amount']
        
        for send in self.sends:
            # Check if amounts match (within 1% to allow for fees)
            time_diff = abs((deposit_time - send['timestamp']).total_seconds() / 3600)
            amount_ratio = min(deposit_amount, send['amount']) / max(deposit_amount, send['amount'])
            
            if time_diff <= window_hours and amount_ratio >= 0.99:
                return send
        
        return None

    def process_sale(self, sale_info: Dict) -> SaleTransaction:
        # Process a sale using the specified accounting method (FIFO, LIFO, HIFO).
        # This is where we:
        # 1. Find available lots to sell from
        # 2. Sort them according to the accounting method
        # 3. Calculate gains/losses for each lot used
        # 4. Track which lots were used and how much remains

        # Get available lots and sort them
        available_lots = [lot for lot in self.lots if lot.remaining > 0]
        
        # Sort lots based on accounting method
        if self.accounting_method == AccountingMethod.FIFO:
            available_lots.sort(key=lambda x: x.timestamp)
            
        elif self.accounting_method == AccountingMethod.LIFO:
            available_lots.sort(key=lambda x: x.timestamp, reverse=True)
            
        elif self.accounting_method == AccountingMethod.HIFO:
            available_lots.sort(key=lambda x: (
                -float(str(x.cost_per_unit) if x.cost_per_unit != 0 else '0'),
                x.timestamp
            ))
            
        elif self.accounting_method == AccountingMethod.LOFO:
            available_lots.sort(key=lambda x: (
                float(str(x.cost_per_unit) if x.cost_per_unit != 0 else '0'),
                x.timestamp
            ))

        amount_to_sell = abs(Decimal(str(sale_info['amount'])))
        
        # Verify we have enough total amount
        total_available = sum(lot.remaining for lot in available_lots)

        if total_available < amount_to_sell:
            log_error(f"Insufficient lots available for {self.symbol} sale:")
            log_error(f"Need: {amount_to_sell}, Have: {total_available}")
            raise ValueError(f"Insufficient lots available for {self.symbol} sale")

        # Process sale using available lots in sorted order
        remaining_to_sell = amount_to_sell
        lot_details = []
        lots_sold = []
        is_long_term = True
        total_cost_basis = Decimal('0')
        total_proceeds = Decimal('0')
        proceeds = Decimal(str(sale_info['proceeds']))

        # Check if this is a multi-lot sale
        needed_lots = len([lot for lot in available_lots if lot.remaining > 0])
        multi_lot_sale = needed_lots > 1

        for i, lot in enumerate(available_lots):
            if remaining_to_sell <= 0:
                break

            amount_from_lot = min(lot.remaining, remaining_to_sell)
            
            if amount_from_lot > 0:
                # Calculate proceeds portion
                if multi_lot_sale:
                    # For any multi-lot sale, distribute proceeds proportionally
                    proceeds_portion = normalize_usd(
                        (amount_from_lot / amount_to_sell) * proceeds
                    )
                else:
                    # Single lot - use all proceeds
                    proceeds_portion = proceeds

                # Update running total of proceeds
                total_proceeds += proceeds_portion

                # Calculate cost basis
                cost_basis_portion = normalize_usd(
                    amount_from_lot * lot.cost_per_unit
                )

                # Calculate gain/loss
                gain_loss = proceeds_portion - cost_basis_portion

                # Check if long term
                holding_period = sale_info['timestamp'] - lot.timestamp
                is_lot_long_term = holding_period.days > 365
                if not is_lot_long_term:
                    is_long_term = False

                # Create lot detail
                lot_detail = SaleLotDetail(
                    lot_id=lot.lot_id,
                    amount_sold=amount_from_lot,
                    cost_basis=cost_basis_portion,
                    proceeds=proceeds_portion,
                    gain_loss=gain_loss,
                    is_long_term=is_lot_long_term,
                    purchase_date=lot.timestamp,
                    holding_period_days=holding_period.days,
                    cost_basis_per_unit=lot.cost_per_unit
                )
                lot_details.append(lot_detail)
                lots_sold.append(lot.lot_id)

                # Update running totals
                total_cost_basis += cost_basis_portion
                remaining_to_sell -= amount_from_lot
                lot.remaining -= amount_from_lot

        # Create sale transaction
        sale = SaleTransaction(
            timestamp=sale_info['timestamp'],
            amount=abs(Decimal(str(sale_info['amount']))),
            proceeds=proceeds,
            cost_basis=total_cost_basis,
            source=sale_info['source'],
            transaction_id=sale_info['transaction_id'],
            lot_details=lot_details,
            is_long_term=is_long_term,
            lots_sold=', '.join(lots_sold)
        )

        self.sales.append(sale)
        return sale

    def get_current_holdings(self) -> Dict:
        # Calculate current holdings and average cost basis
        # Returns a dictionary with total amounts, cost basis, and other metrics
        
        # Filter out Pro Withdrawals and Pro Deposits when calculating holdings
        valid_lots = [
            lot for lot in self.lots 
            if lot.transaction_type not in [TransactionType.PRO_WITHDRAWAL, TransactionType.PRO_DEPOSIT]
        ]
        
        # Basic calculations
        total_amount = sum(lot.remaining for lot in valid_lots)
        total_cost_basis = sum(lot.remaining * lot.cost_per_unit for lot in valid_lots)
        total_fees = sum(lot.cost_basis - (lot.amount * lot.cost_per_unit) for lot in valid_lots)
        
        # Get all lots with remaining amounts
        lots_with_balance = [lot for lot in valid_lots if lot.remaining > 0]
        
        # Calculate average cost basis
        avg_cost_basis = (
            (total_cost_basis / total_amount) if total_amount > 0 
            else Decimal('0')
        )
        
        # Get current price for unrealized P/L calculation
        current_price = get_historical_price(self.symbol, datetime.now())
        current_value = total_amount * Decimal(str(current_price))
        unrealized_pl = current_value - total_cost_basis
        unrealized_pl_pct = (
            (current_value / total_cost_basis - 1)
            if total_cost_basis > 0 else Decimal('0')
        )
        
        # Calculate exchange distribution
        exchange_amounts = defaultdict(Decimal)
        for lot in valid_lots:
            if lot.remaining > 0:
                if 'external' in lot.source.lower() or 'wallet' in lot.source.lower():
                    exchange_amounts['External Wallet(s)'] += lot.remaining
                else:
                    exchange_amounts[lot.source] += lot.remaining
                    
        distribution = []
        for exchange, amount in exchange_amounts.items():
            percentage = (amount / total_amount * 100).quantize(Decimal('0.1'))
            distribution.append(f"{exchange}: {percentage}%")
        exchange_distribution = ", ".join(distribution)
        
        # Calculate holding periods
        now = datetime.now()
        active_lots = [lot for lot in valid_lots if lot.remaining > 0]
        if active_lots:
            first_tx_date = min(lot.timestamp for lot in active_lots)
            last_tx_date = max(lot.timestamp for lot in active_lots)
            days_held = (now - first_tx_date).days
            
            # Calculate average hold time
            weighted_days = sum(
                lot.remaining * (now - lot.timestamp).days 
                for lot in active_lots
            )
            avg_hold_time = int(weighted_days / total_amount) if total_amount > 0 else 0
            
            # Calculate long/short term holdings
            long_term_amount = sum(
                lot.remaining 
                for lot in active_lots 
                if (now - lot.timestamp).days > 365
            )
            short_term_amount = total_amount - long_term_amount
            
            # Find price extremes
            lot_prices = [lot.cost_per_unit for lot in active_lots if lot.cost_per_unit > 0]
            lowest_price = min(lot_prices) if lot_prices else Decimal('0')
            highest_price = max(lot_prices) if lot_prices else Decimal('0')
        else:
            last_tx_date = None
            days_held = 0
            avg_hold_time = 0
            long_term_amount = Decimal('0')
            short_term_amount = Decimal('0')
            lowest_price = Decimal('0')
            highest_price = Decimal('0')
        
        today = datetime.now()
        spot_price_key = f'spot_price_{today.strftime("%Y-%m-%d")}'
        total_usd_key = f'total_usd_{today.strftime("%Y-%m-%d")}'
        
        return {
            'symbol': self.symbol,
            'total_amount': total_amount,
            'total_cost_basis': total_cost_basis,
            'average_cost_basis': avg_cost_basis,
            'total_fees': total_fees,
            spot_price_key: Decimal(str(current_price)),
            total_usd_key: current_value,
            'unrealized_pl': unrealized_pl,
            'unrealized_pl_pct': unrealized_pl_pct,
            'exchange_distribution': exchange_distribution,
            'lot_count': len(lots_with_balance),  # Use actual count of lots with balance
            'last_tx_date': last_tx_date,
            'days_held': days_held,
            'avg_hold_time': avg_hold_time,
            'long_term_amount': long_term_amount,
            'short_term_amount': short_term_amount,
            'lowest_price': lowest_price,
            'highest_price': highest_price,
            'lots': lots_with_balance  # Add all lots with remaining balance
        }

    def process_deposit_with_send(self, deposit: Dict, matching_send: Dict) -> None:
        # Process a deposit that matches a previous send
        # Create new lot(s) with the Total USD as cost basis

        lot = AssetLot(
            timestamp=deposit['timestamp'],
            amount=deposit['amount'],
            cost_basis=deposit['total_usd'],  # Use Total USD from deposit
            source=deposit['source'],
            transaction_type=TransactionType.DEPOSIT,
            remaining=deposit['amount'],
            transaction_id=f"{deposit['transaction_id']}_from_{matching_send['transaction_id']}",
            lot_id=deposit['lot_id']
        )
        self.add_lot(lot)
        
        # Track the transfer with Total USD as cost basis
        self.transfers.append({
            'from_exchange': matching_send['source'],
            'to_exchange': deposit['source'],
            'timestamp': deposit['timestamp'],
            'amount': deposit['amount'],
            'cost_basis': deposit['total_usd'],  # Use Total USD from deposit
            'from_tx_id': matching_send['transaction_id'],
            'to_tx_id': deposit['transaction_id']
        })

    def validate_lot_selection(self, lots_used: List[Dict], method: AccountingMethod) -> bool:
        # Validate that lots were selected according to the specified method

        if not lots_used:
            return True
            
        # Convert to list of tuples for easier comparison
        lot_info = [(lot['timestamp'], Decimal(str(lot['basis_per_unit']))) for lot in lots_used]
        
        # Check if lots are in correct order based on method
        for i in range(len(lot_info) - 1):
            current, next_lot = lot_info[i], lot_info[i + 1]
            
            if method == AccountingMethod.FIFO:
                if current[0] > next_lot[0]:  # Should be chronological
                    return False
            elif method == AccountingMethod.LIFO:
                if current[0] < next_lot[0]:  # Should be reverse chronological
                    return False
            elif method == AccountingMethod.HIFO:
                if current[1] < next_lot[1]:  # Should be highest cost first
                    return False
            elif method == AccountingMethod.LOFO:
                if current[1] > next_lot[1]:  # Should be lowest cost first
                    return False
                    
        return True

    def validate_lot_selection_strict(self, lots_used: List[Dict], method: AccountingMethod) -> None:
        # Strictly validate lot selection matches accounting method

        if not lots_used:
            return
            
        if method == AccountingMethod.FIFO:
            # Verify strictly chronological order
            timestamps = [lot['timestamp'] for lot in lots_used]
            if timestamps != sorted(timestamps):
                raise ValueError("FIFO violation: Lots not used in chronological order")
                
        elif method == AccountingMethod.LIFO:
            # Verify strictly reverse chronological order
            timestamps = [lot['timestamp'] for lot in lots_used]
            if timestamps != sorted(timestamps, reverse=True):
                raise ValueError("LIFO violation: Lots not used in reverse chronological order")
                
        elif method == AccountingMethod.HIFO:
            # Verify strictly descending cost basis
            costs = [Decimal(str(lot['basis_per_unit'])) for lot in lots_used]
            if costs != sorted(costs, reverse=True):
                raise ValueError("HIFO violation: Lots not used in descending cost order")
                
        elif method == AccountingMethod.LOFO:
            # Verify strictly ascending cost basis
            costs = [Decimal(str(lot['basis_per_unit'])) for lot in lots_used]
            if costs != sorted(costs):
                raise ValueError("LOFO violation: Lots not used in ascending cost order")

    def verify_cost_basis(self, lot_details: List[SaleLotDetail], total_cost_basis: Decimal) -> bool:
        # Verify cost basis calculations are correct

        calculated_total = sum(detail.cost_basis for detail in lot_details)
        difference = abs(calculated_total - total_cost_basis)
        
        # Allow for tiny rounding differences (less than 1 cent)
        if difference > Decimal('0.01'):
            log_error(f"Cost basis mismatch: {calculated_total} vs {total_cost_basis}")
            return False
        return True

    def verify_lot_exhaustion(self, amount_to_sell: Decimal, lots_used: List[Dict]) -> bool:
        # Verify that lots were properly exhausted

        total_amount_used = sum(lot['amount'] for lot in lots_used)
        difference = abs(amount_to_sell - total_amount_used)
        
        # Allow for tiny rounding differences
        if difference > Decimal('0.000000000000000001'):
            log_error(f"Amount mismatch: needed {amount_to_sell}, used {total_amount_used}")
            return False
        return True

    def validate_transaction_amounts(self, sale_info: Dict, used_lots: List[Dict]) -> None:
        # Ensure transaction amounts balance correctly

        sale_amount = abs(Decimal(str(sale_info['amount'])))
        total_lot_amount = sum(Decimal(str(lot['amount'])) for lot in used_lots)
        
        if abs(sale_amount - total_lot_amount) > Decimal('0.000000000000000001'):
            error_msg = (f"Amount mismatch in {self.symbol} sale: "
                        f"Sale amount {sale_amount} != Total lot amount {total_lot_amount}")
            log_error(error_msg)
            raise ValueError(error_msg)

    def validate_cost_basis_calculation(self, lot_details: List[SaleLotDetail]) -> None:
        # Ensure cost basis is calculated correctly for each lot

        for detail in lot_details:
            calculated = detail.amount_sold * detail.cost_basis_per_unit
            if abs(calculated - detail.cost_basis) > Decimal('0.01'):
                error_msg = (f"Cost basis calculation error in lot {detail.lot_id}: "
                            f"Calculated {calculated} != Recorded {detail.cost_basis}")
                log_error(error_msg)
                raise ValueError(error_msg)

    def validate_proceeds_calculation(self, sale_info: Dict, lot_details: List[SaleLotDetail]) -> None:
        # Ensure proceeds are distributed correctly across lots

        total_proceeds = sum(detail.proceeds for detail in lot_details)
        sale_proceeds = Decimal(str(sale_info['proceeds']))
        
        if abs(total_proceeds - sale_proceeds) > Decimal('0.01'):
            error_msg = (f"Proceeds distribution error: "
                        f"Total proceeds {total_proceeds} != Sale proceeds {sale_proceeds}")
            log_error(error_msg)
            raise ValueError(error_msg)

    def verify_accounting_method_consistency(self, used_lots: List[Dict]) -> None:
        # Verify lots were selected according to the specified accounting method

        if len(used_lots) <= 1:
            return
            
        for i in range(len(used_lots) - 1):
            current, next_lot = used_lots[i], used_lots[i + 1]
            
            if self.accounting_method == AccountingMethod.FIFO:
                if current['timestamp'] > next_lot['timestamp']:
                    raise ValueError(f"FIFO order violation in {self.symbol} sale")
                    
            elif self.accounting_method == AccountingMethod.LIFO:
                if current['timestamp'] < next_lot['timestamp']:
                    raise ValueError(f"LIFO order violation in {self.symbol} sale")
                    
            elif self.accounting_method == AccountingMethod.HIFO:
                if Decimal(str(current['basis_per_unit'])) < Decimal(str(next_lot['basis_per_unit'])):
                    raise ValueError(f"HIFO order violation in {self.symbol} sale")
                    
            elif self.accounting_method == AccountingMethod.LOFO:
                if Decimal(str(current['basis_per_unit'])) > Decimal(str(next_lot['basis_per_unit'])):
                    raise ValueError(f"LOFO order violation in {self.symbol} sale")

class TransactionProcessor:
    # Process all transactions and manage asset trackers
    # This class:
    # 1. Keeps track of all purchases (lots)
    # 2. Records all sales and their details
    # 3. Tracks staking income and gifts
    # 4. Handles transfers between exchanges

    def __init__(self, accounting_method: AccountingMethod = AccountingMethod.HIFO):
        self.asset_trackers: Dict[str, AssetTracker] = {}
        self.current_prices: Dict[str, Decimal] = {}
        self.accounting_method = accounting_method
        self.all_errors = []            # Collect all errors across assets
        self.proceeds_mismatches = []   # Collect proceeds mismatches
        self.mismatch_details = []      # Store detailed mismatch info

    def get_tracker(self, symbol: str) -> AssetTracker:
        # Get or create asset tracker for symbol

        if symbol not in self.asset_trackers:
            self.asset_trackers[symbol] = AssetTracker(symbol, self.accounting_method)
        return self.asset_trackers[symbol]
    
    def process_transaction(self, row: pd.Series) -> None:
        # Process a single transaction

        try:
            # Skip USD deposits/withdrawals, Exchange Withdrawals, and Pro Withdrawals
            if (row['Asset'] == 'USD' and row['Type'].lower() in ['deposit', 'withdrawal']) or \
               'exchange withdrawal' in row['Type'].lower() or \
               'pro withdrawal' in row['Type'].lower():
                return
            
            # Get transaction type
            tx_type = TransactionType.from_string(row['Type'])
            if tx_type is None:
                return
                
            # Get or create tracker for this asset
            tracker = self.get_tracker(row['Asset'])
            
            # Process transaction based on type
            if tx_type in [TransactionType.BUY, TransactionType.CONVERT]:
                # Handle purchases
                lot = AssetLot(
                    timestamp=row['Timestamp'],
                    amount=abs(Decimal(str(row['Amount']))),
                    cost_basis=abs(Decimal(str(row['Total USD']))),
                    source=row['Source'],
                    transaction_type=tx_type,
                    remaining=abs(Decimal(str(row['Amount']))),
                    transaction_id=row['ID'],
                    lot_id=row['Lot ID']
                )
                tracker.add_lot(lot)
            
            elif tx_type == TransactionType.SELL:
                # Handle sales
                sale_info = {
                    'timestamp': row['Timestamp'],
                    'amount': abs(Decimal(str(row['Amount']))),
                    'proceeds': abs(Decimal(str(row['Subtotal']))),
                    'source': row['Source'],
                    'transaction_id': row['ID']
                }
                tracker.process_sale(sale_info)
                
            elif tx_type == TransactionType.STAKE:
                # Handle staking income
                staking = StakingRecord(
                    timestamp=row['Timestamp'],
                    amount=abs(Decimal(str(row['Amount']))),
                    market_value=abs(Decimal(str(row['Total USD']))),
                    source=row['Source'],
                    transaction_id=row['ID']
                )
                tracker.staking_income.append(staking)
                
                # Add as lot with zero cost basis
                lot = AssetLot(
                    timestamp=row['Timestamp'],
                    amount=abs(Decimal(str(row['Amount']))),
                    cost_basis=Decimal('0'),
                    source=row['Source'],
                    transaction_type=tx_type,
                    remaining=abs(Decimal(str(row['Amount']))),
                    transaction_id=row['ID'],
                    lot_id=row['Lot ID']
                )
                tracker.add_lot(lot)
                
            elif tx_type == TransactionType.GIFT:
                # Handle gifts received
                gift = GiftRecord(
                    timestamp=row['Timestamp'],
                    amount=abs(Decimal(str(row['Amount']))),
                    market_value=abs(Decimal(str(row['Total USD']))),
                    source=row['Source'],
                    transaction_id=row['ID'],
                    gift_type=row['Type']
                )
                tracker.gifts.append(gift)
                
                # Add as lot with market value for cost basis tracking
                lot = AssetLot(
                    timestamp=row['Timestamp'],
                    amount=abs(Decimal(str(row['Amount']))),
                    cost_basis=Decimal('0'),  # Original cost basis is 0
                    source=row['Source'],
                    transaction_type=tx_type,
                    remaining=abs(Decimal(str(row['Amount']))),
                    transaction_id=row['ID'],
                    lot_id=row['Lot ID'],
                    gift_market_value=abs(Decimal(str(row['Total USD']))),  # Store market value for future sales
                )
                tracker.add_lot(lot)
                
            elif tx_type == TransactionType.SEND:
                # Track send with Total USD as cost basis
                send_info = {
                    'timestamp': row['Timestamp'],
                    'amount': abs(Decimal(str(row['Amount']))),
                    'source': row['Source'],
                    'transaction_id': row['ID'],
                    'cost_basis': abs(Decimal(str(row['Total USD']))),  # Use Total USD from input file
                    'lots_info': []  # Keep this for compatibility
                }
                tracker.sends.append(send_info)
                
            elif tx_type == TransactionType.DEPOSIT:
                # Check for matching send transaction
                deposit_info = {
                    'timestamp': row['Timestamp'],
                    'amount': abs(Decimal(str(row['Amount']))),
                    'source': row['Source'],
                    'transaction_id': row['ID'],
                    'lot_id': row['Lot ID'],
                    'total_usd': abs(Decimal(str(row['Total USD']))),  # Add this line to include Total USD
                }
                
                matching_send = tracker.find_matching_send(deposit_info)
                if matching_send:
                    tracker.process_deposit_with_send(deposit_info, matching_send)
                else:
                    # Create new lot for unmatched deposit
                    lot = AssetLot(
                        timestamp=row['Timestamp'],
                        amount=abs(Decimal(str(row['Amount']))),
                        cost_basis=abs(Decimal(str(row['Total USD']))),  # Use Total USD as cost basis
                        source=row['Source'],
                        transaction_type=tx_type,
                        remaining=abs(Decimal(str(row['Amount']))),
                        transaction_id=row['ID'],
                        lot_id=row['Lot ID']
                    )
                    tracker.add_lot(lot)
            
            elif tx_type == TransactionType.ADMIN_DEBIT:
                # Remove from holdings without creating a sale
                amount_to_remove = abs(Decimal(str(row['Amount'])))
                for lot in sorted(tracker.lots, key=lambda x: x.cost_per_unit):
                    if amount_to_remove <= 0:
                        break
                    remove_from_lot = min(lot.remaining, amount_to_remove)
                    lot.remaining -= remove_from_lot
                    amount_to_remove -= remove_from_lot
                
            elif tx_type == TransactionType.CONVERT:
                # Handle the sale part of conversion - use absolute value of Subtotal for proceeds
                sale_info = {
                    'timestamp': row['Timestamp'],
                    'amount': abs(Decimal(str(row['Amount']))),
                    'proceeds': abs(Decimal(str(row['Subtotal']))),  # Added abs()
                    'source': row['Source'],
                    'transaction_id': row['ID'] + '_SELL'
                }
                tracker.process_sale(sale_info)
                
                # Handle the buy part of conversion
                # Note: You'll need to get the received asset details from the Notes or related transaction

                received_asset = extract_received_asset(row['Notes'])  # You'll need to implement this
                if received_asset:
                    received_tracker = self.get_tracker(received_asset['symbol'])
                    lot = AssetLot(
                        timestamp=row['Timestamp'],
                        amount=received_asset['amount'],
                        cost_basis=abs(Decimal(str(row['Total USD']))),  # Use the same USD value
                        source=row['Source'],
                        transaction_type=tx_type,
                        remaining=received_asset['amount'],
                        transaction_id=row['ID'] + '_BUY',
                        lot_id=row['Lot ID']
                    )
                    received_tracker.add_lot(lot)
            
            elif tx_type == TransactionType.DIVIDEND:
                # Handle dividend as new asset with zero cost basis
                lot = AssetLot(
                    timestamp=row['Timestamp'],
                    amount=abs(Decimal(str(row['Amount']))),
                    cost_basis=Decimal('0'),  # Always zero cost basis for dividends
                    source=row['Source'],
                    transaction_type=tx_type,
                    remaining=abs(Decimal(str(row['Amount']))),
                    transaction_id=row['ID'],
                    lot_id=row['Lot ID']
                )
                tracker.add_lot(lot)
            
            elif tx_type == TransactionType.WITHDRAWAL:
                # Track withdrawal with Total USD as cost basis
                tracker.withdrawals.append({
                    'timestamp': row['Timestamp'],
                    'amount': abs(Decimal(str(row['Amount']))),
                    'source': row['Source'],
                    'transaction_id': row['ID'],
                    'cost_basis': abs(Decimal(str(row['Total USD']))),  # Use Total USD from input file
                })
            
        except InsufficientLotsError as e:
            # Add the error here instead
            self.all_errors.append(e.args[0])
        except Exception as e:
            log_error(f"Error processing transaction: {row['ID']}")
            log_error(f"Error details: {str(e)}\n")
            # Log any unexpected errors
            self.all_errors.append({
                'timestamp': row['Timestamp'],
                'asset': row['Asset'],
                'error': str(e),
                'transaction': row.to_dict()
            })

    def process_all_transactions(self, df: pd.DataFrame) -> None:
        # Process all transactions in chronological order
 
        # Sort by timestamp
        df = df.sort_values('Timestamp')
        
        # Calculate gains first to get lot assignments
        gains_df, validation_errors = calculate_gains(df, self.accounting_method)
        
        # Add validation errors to all_errors
        self.all_errors.extend(validation_errors)
        
        # Process each transaction
        for _, row in df.iterrows():
            try:
                self.process_transaction(row)
            except Exception as e:
                # Log any unexpected errors
                self.all_errors.append({
                    'timestamp': row['Timestamp'],
                    'asset': row['Asset'],
                    'error': str(e),
                    'transaction': row.to_dict()
                })

    def write_results(self, output_file='crypto_gains.xlsx'):
        # Create a workbook with formatting capabilities and write all results
        # Includes transaction details, error summaries, and proceeds mismatches

        workbook = openpyxl.Workbook()
        
        # Write transactions worksheet
        ws = workbook.active
        ws.title = 'Transactions'
        
        # Add headers
        headers = ['Timestamp', 'Asset', 'Type', 'Amount', 'Cost Basis', 'Proceeds', 
                  'Gain/Loss', 'Holding Period', 'Term', 'Notes']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Define formats
        error_fill = openpyxl.styles.PatternFill(
            start_color='FFE6E6',  # Light red
            end_color='FFE6E6',
            fill_type='solid'
        )
        
        # Write transaction data
        row_num = 2
        for asset in self.asset_trackers.values():
            for sale in asset.processed_sales:
                # Write the row data
                ws.cell(row=row_num, column=1, value=sale['timestamp'])
                ws.cell(row=row_num, column=2, value=asset.symbol)
                # ... other columns ...
                
                # If this sale had an error, highlight the row
                if sale.get('requires_attention'):
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = error_fill
                
                row_num += 1
        
        # Add error summary worksheet
        if self.all_errors:
            error_ws = workbook.create_sheet('Errors')
            error_ws.append(['Timestamp', 'Asset', 'Error', 'Details'])
            
            for error in self.all_errors:
                error_ws.append([
                    error['timestamp'],
                    error['asset'],
                    error['error'],
                    str(error['transaction'])
                ])
            
            # Auto-adjust column widths
            for column in error_ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                error_ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(output_file)
        
        # Print error summary to console
        if self.all_errors:
            print("\nERRORS FOUND:")
            print("=============")
            for error in self.all_errors:
                print(f"\nAsset: {error['asset']}")
                print(f"Timestamp: {error['timestamp']}")
                print(f"Error: {error['error']}")
                print("-" * 50)

    def add_proceeds_mismatch(self, symbol: str, sale_timestamp, expected: Decimal, actual: Decimal):
        # Track a proceeds mismatch
        self.proceeds_mismatches.append({
            'symbol': symbol,
            'timestamp': sale_timestamp,
            'expected': expected,
            'actual': actual,
            'difference': abs(expected - actual)
        })

    def add_mismatch_details(self, symbol: str, details: Dict) -> None:
        # Store detailed information about a proceeds mismatch

        details['symbol'] = symbol
        self.mismatch_details.append(details)
    
    def get_mismatch_details(self) -> List[Dict]:
        # Retrieve the stored mismatch details
        return self.mismatch_details

class GainsCalculator:
    # Calculate and generate reports for crypto gains/losses

    def __init__(self, processor: TransactionProcessor):
        self.processor = processor
    
    def calculate_yearly_summary(self) -> Dict[str, Dict]:
        # Calculate yearly summary of gains/losses, staking income, and gifts.
        # Returns a dictionary organized by year, then by cryptocurrency symbol.
        # Each symbol contains lists of short-term sales, long-term sales,
        # staking income, and gifts for that year.

        yearly_data = {}
        staking_lots_sold = set()  # Track which staking lots were sold in their receive year
        gift_lots_sold = set()     # Track which gift lots were sold in their receive year
        
        # First pass: Track which lots are sold in their receive year
        for symbol, tracker in self.processor.asset_trackers.items():
            for sale in tracker.sales:
                sale_year = sale.timestamp.year
                for lot_detail in sale.lot_details:
                    # Find original lot
                    for lot in tracker.lots:
                        if lot.lot_id == lot_detail.lot_id:
                            if lot.transaction_type == TransactionType.STAKE and lot.timestamp.year == sale_year:
                                staking_lots_sold.add((symbol, lot.lot_id))
                            elif lot.transaction_type == TransactionType.GIFT and lot.timestamp.year == sale_year:
                                gift_lots_sold.add((symbol, lot.lot_id))
        
        # Second pass: Calculate summaries
        for symbol, tracker in self.processor.asset_trackers.items():
            # Process sales first
            for sale in tracker.sales:
                year = sale.timestamp.year
                if year not in yearly_data:
                    yearly_data[year] = {}
                if symbol not in yearly_data[year]:
                    yearly_data[year][symbol] = {
                        'short_term_sales': [],
                        'long_term_sales': [],
                        'staking_income': [],
                        'gifts': []
                    }
                
                if sale.is_long_term:
                    yearly_data[year][symbol]['long_term_sales'].append(sale)
                else:
                    yearly_data[year][symbol]['short_term_sales'].append(sale)
            
            # Process staking income
            for staking in tracker.staking_income:
                year = staking.timestamp.year
                # Only include staking income if it wasn't sold in the same year
                if (symbol, staking.transaction_id) not in staking_lots_sold:
                    if year not in yearly_data:
                        yearly_data[year] = {}
                    if symbol not in yearly_data[year]:
                        yearly_data[year][symbol] = {
                            'short_term_sales': [],
                            'long_term_sales': [],
                            'staking_income': [],
                            'gifts': []
                        }
                    yearly_data[year][symbol]['staking_income'].append(staking)
            
            # Process gifts (Learning Rewards and Receives)
            for gift in tracker.gifts:
                year = gift.timestamp.year
                # Only include gift if it wasn't sold in the same year
                if (symbol, gift.transaction_id) not in gift_lots_sold:
                    if year not in yearly_data:
                        yearly_data[year] = {}
                    if symbol not in yearly_data[year]:
                        yearly_data[year][symbol] = {
                            'short_term_sales': [],
                            'long_term_sales': [],
                            'staking_income': [],
                            'gifts': []
                        }
                    yearly_data[year][symbol]['gifts'].append(gift)
        
        return yearly_data

def generate_excel_report(yearly_data: Dict, processor: TransactionProcessor, suffix: str = '') -> None:
    #Generate Excel report with yearly summaries and current holdings

    try:
        with pd.ExcelWriter(f'ALL-crypto-profit-and-loss{suffix}.xlsx', engine='openpyxl') as writer:
            # Add accounting method info sheet
            method_info = pd.DataFrame([{
                'Accounting Method': processor.accounting_method.value,
                'Processing Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Warning': 'This method must be used consistently across tax years'
            }])
            method_info.to_excel(writer, sheet_name='Method Info', index=False)
            format_excel_worksheet(writer.sheets['Method Info'], method_info, 'Method Info')  # Changed
            
            # Generate yearly sheets
            for year, year_data in sorted(yearly_data.items()):
                rows = []
                for symbol, data in year_data.items():
                    # Short-term sales
                    if data['short_term_sales']:
                        dates = [sale.timestamp for sale in data['short_term_sales']]
                        date_range = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}"
                        proceeds = sum(sale.proceeds for sale in data['short_term_sales'])
                        cost = sum(sale.cost_basis for sale in data['short_term_sales'])
                        rows.append({
                            'Asset': symbol,
                            'Category': 'Short-term Sales',
                            'Total Amount': sum(sale.amount for sale in data['short_term_sales']),
                            'Proceeds': proceeds,
                            'Cost Basis': cost,
                            'Gain/Loss': proceeds - cost,
                            'Transaction Count': len(data['short_term_sales']),
                            'Date Range': date_range
                        })
                    
                    # Long-term sales
                    if data['long_term_sales']:
                        dates = [sale.timestamp for sale in data['long_term_sales']]
                        date_range = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}"
                        proceeds = sum(sale.proceeds for sale in data['long_term_sales'])
                        cost = sum(sale.cost_basis for sale in data['long_term_sales'])
                        rows.append({
                            'Asset': symbol,
                            'Category': 'Long-term Sales',
                            'Total Amount': sum(sale.amount for sale in data['long_term_sales']),
                            'Proceeds': proceeds,
                            'Cost Basis': cost,
                            'Gain/Loss': proceeds - cost,
                            'Transaction Count': len(data['long_term_sales']),
                            'Date Range': date_range
                        })
                    
                    # Staking income
                    if data['staking_income']:
                        dates = [stake.timestamp for stake in data['staking_income']]
                        date_range = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}"
                        market_value = sum(stake.market_value for stake in data['staking_income'])
                        rows.append({
                            'Asset': symbol,
                            'Category': 'Staking Income',
                            'Total Amount': sum(stake.amount for stake in data['staking_income']),
                            'Proceeds': market_value,
                            'Cost Basis': 0,
                            'Gain/Loss': market_value,
                            'Transaction Count': len(data['staking_income']),
                            'Date Range': date_range
                        })
                    
                    # Gifts received
                    if data['gifts']:
                        dates = [gift.timestamp for gift in data['gifts']]
                        date_range = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}"
                        market_value = sum(gift.market_value for gift in data['gifts'])
                        rows.append({
                            'Asset': symbol,
                            'Category': 'Gifts Received',
                            'Total Amount': sum(gift.amount for gift in data['gifts']),
                            'Proceeds': market_value,  # Market value at time of receipt
                            'Cost Basis': 0,  # Cost basis is 0 for gifts
                            'Gain/Loss': market_value,  # Full market value is the gain
                            'Transaction Count': len(data['gifts']),
                            'Date Range': date_range
                        })
                
                if rows:
                    df = pd.DataFrame(rows)
                    df = df.sort_values(['Category', 'Asset'])
                    
                    # Convert numeric columns to float
                    numeric_columns = [
                        'Total Amount', 'Proceeds', 'Cost Basis', 'Gain/Loss', 
                        'Transaction Count'
                    ]
                    for col in numeric_columns:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                    
                    df.to_excel(writer, sheet_name=f'Year_{year}', index=False)
                    format_excel_worksheet(writer.sheets[f'Year_{year}'], df, f'Year_{year}')  # Changed
            
            # Generate current holdings sheet
            holdings_rows = []
            for symbol, tracker in processor.asset_trackers.items():
                holdings = tracker.get_current_holdings()
                if holdings['total_amount'] > 0:
                    lots_remaining = []
                    for lot in holdings['lots']:
                        # Format amount with stripped trailing zeros
                        amount_str = strip_trailing_zeros(lot.remaining)
                        # Add markdown-style bold markers around lot ID and parentheses
                        lots_remaining.append(f"{lot.lot_id} ({amount_str})")
                    # Sort lots by lot_id
                    lots_remaining.sort()
                    holdings['lots_remaining'] = ', '.join(lots_remaining)
                    holdings_rows.append(holdings)

            
            if holdings_rows:
                holdings_df = pd.DataFrame(holdings_rows)
                holdings_df = format_holdings_dataframe(holdings_df)
                holdings_df.to_excel(writer, sheet_name='Current Holdings', index=False)
                format_excel_worksheet(writer.sheets['Current Holdings'], holdings_df, 'Current Holdings')
            
            # Add transfers sheet
            transfer_rows = []
            for symbol, tracker in processor.asset_trackers.items():
                # Add exchange-to-exchange transfers
                for transfer in tracker.transfers:
                    transfer_rows.append({
                        'Asset': symbol,
                        'Timestamp': transfer['timestamp'],
                        'From Exchange': transfer['from_exchange'],
                        'To Exchange': transfer['to_exchange'],
                        'Amount': abs(transfer['amount']),
                        'Cost Basis': transfer['cost_basis'],
                        'From Transaction ID': transfer['from_tx_id'],
                        'To Transaction ID': transfer['to_tx_id'],
                        'Type': 'Exchange Transfer'
                    })
                
                # Add sends and withdrawals
                for send in tracker.sends:
                    if symbol != 'USD':  # Skip USD transactions
                        transfer_rows.append({
                            'Asset': symbol,
                            'Timestamp': send['timestamp'],
                            'From Exchange': send['source'],
                            'To Exchange': 'External Wallet',
                            'Amount': abs(send['amount']),
                            'Cost Basis': sum(lot['cost_basis'] for lot in send['lots_info']),
                            'From Transaction ID': send['transaction_id'],
                            'To Transaction ID': '',
                            'Type': 'Send'
                        })
                
                # Add withdrawals that aren't already tracked as sends
                for withdrawal in tracker.withdrawals:
                    if symbol != 'USD':  # Skip USD transactions
                        transfer_rows.append({
                            'Asset': symbol,
                            'Timestamp': withdrawal['timestamp'],
                            'From Exchange': withdrawal['source'],
                            'To Exchange': 'External Wallet',
                            'Amount': abs(withdrawal['amount']),
                            'Cost Basis': withdrawal['cost_basis'],
                            'From Transaction ID': withdrawal['transaction_id'],
                            'To Transaction ID': '',
                            'Type': 'Withdrawal'
                        })
            
            if transfer_rows:
                transfer_df = pd.DataFrame(transfer_rows)
                
                # Convert numeric columns to float
                numeric_columns = ['Amount', 'Cost Basis']
                for col in numeric_columns:
                    if col in transfer_df.columns:
                        transfer_df[col] = pd.to_numeric(transfer_df[col], errors='coerce')
                
                transfer_df.sort_values('Timestamp', inplace=True)
                transfer_df.to_excel(writer, sheet_name='Transfers', index=False)
                format_excel_worksheet(writer.sheets['Transfers'], transfer_df, 'Transfers')  # Changed
            
            # Add Sale Details sheet
            sale_detail_rows = []
            
            # First, process the yearly data to get proper term assignments
            sales_with_terms = {}
            for year, assets in yearly_data.items():
                for symbol, data in assets.items():
                    # Track which sales are short vs long term
                    for sale in data['short_term_sales']:
                        sales_with_terms[sale.transaction_id] = 'Short'
                    for sale in data['long_term_sales']:
                        sales_with_terms[sale.transaction_id] = 'Long'

            # Now process sales by timestamp
            for symbol, tracker in processor.asset_trackers.items():
                # Group sales by timestamp
                sales_by_timestamp = {}
                for sale in tracker.sales:
                    timestamp = sale.timestamp
                    if timestamp not in sales_by_timestamp:
                        sales_by_timestamp[timestamp] = []
                    sales_by_timestamp[timestamp].append(sale)
                
                # Process each timestamp group
                for timestamp, sales in sales_by_timestamp.items():
                    # Combine sales with same timestamp
                    total_amount = sum(sale.amount for sale in sales)
                    total_proceeds = sum(sale.proceeds for sale in sales)
                    total_cost_basis = sum(sale.cost_basis for sale in sales)
                    total_gain_loss = total_proceeds - total_cost_basis
                    
                    # Collect all lot details
                    lot_details_map = {}
                    sale_terms = set()  # Track terms for all sales in this group
                    
                    for sale in sales:
                        # Get the correct term from our yearly data mapping
                        sale_terms.add(sales_with_terms.get(sale.transaction_id, 'Mixed'))
                        
                        for lot_detail in sale.lot_details:
                            lot_id = lot_detail.lot_id
                            if lot_id not in lot_details_map:
                                lot_details_map[lot_id] = {
                                    'amount_sold': Decimal('0'),
                                    'proceeds': Decimal('0'),
                                    'cost_basis': Decimal('0'),
                                    'purchase_date': lot_detail.purchase_date,
                                    'is_long_term': lot_detail.is_long_term
                                }
                            lot_details_map[lot_id]['amount_sold'] += lot_detail.amount_sold
                            lot_details_map[lot_id]['proceeds'] += lot_detail.proceeds
                            lot_details_map[lot_id]['cost_basis'] += lot_detail.cost_basis
                    
                    # Create row with lot details
                    lot_ids = sorted(lot_details_map.keys())
                    lot_info = [
                        f"{lot_id} ({lot_details_map[lot_id]['amount_sold']:.8f})"
                        for lot_id in lot_ids
                    ]
                    
                    # Determine term based on the sales in this group
                    if len(sale_terms) == 1:
                        term = sale_terms.pop()  # Use the single term
                    else:
                        term = 'Mixed'  # Only use Mixed if we truly have different terms
                    
                    sale_detail_rows.append({
                        'Asset': symbol,
                        'Sale Date': timestamp,
                        'Amount Sold': total_amount,
                        'Proceeds': total_proceeds,
                        'Cost Basis': total_cost_basis,
                        'Gain/Loss': total_gain_loss,
                        'Source': sales[0].source,
                        'Lots Used': ', '.join(lot_info),
                        'Term': term
                    })

            # Create Sale Details sheet if we have rows
            if sale_detail_rows:
                sale_detail_df = pd.DataFrame(sale_detail_rows)
                
                # Convert numeric columns
                numeric_columns = ['Amount Sold', 'Proceeds', 'Cost Basis', 'Gain/Loss']
                for col in numeric_columns:
                    if col in sale_detail_df.columns:
                        sale_detail_df[col] = pd.to_numeric(sale_detail_df[col], errors='coerce')
                
                # Sort by date and asset
                sale_detail_df.sort_values(['Sale Date', 'Asset'], inplace=True)
                sale_detail_df.to_excel(writer, sheet_name='Sale Details', index=False)
                format_excel_worksheet(writer.sheets['Sale Details'], sale_detail_df, 'Sale Details')  # Changed

            # Ensure at least one sheet exists
            if not writer.sheets:
                empty_df = pd.DataFrame()
                empty_df.to_excel(writer, sheet_name='No Transactions', index=False)
                format_excel_worksheet(writer.sheets['No Transactions'], empty_df, 'No Transactions')  # Changed

            print("\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            print("Report generated successfully!!")
            print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")
            
            print("File SAVED AS...")
            print("    ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            print(f"    ALL-crypto-profit-and-loss{suffix}.xlsx")
            print("    ~~~~~~~~~~~~~~~~~~~~~~~~~~~****~~~~~\n")
        
    except Exception as e:
        log_error(f"Error generating Excel report: {str(e)}")
        raise

def extract_received_asset(notes: str) -> Optional[Dict]:
    # Extract received asset information from conversion notes
    # Common patterns:
    # "Converted 1.5 BTC to 25.3 ETH"
    # "Convert from 1.5 BTC to 25.3 ETH"
    # "1.5 BTC converted to 25.3 ETH"
    # "Conversion: 1.5 BTC → 25.3 ETH"    

    try:
        if not notes or not isinstance(notes, str):
            return None
            
        notes = notes.lower().strip()
        
        patterns = [
            r'(?:converted|convert|conversion).*?(\d+\.?\d*)\s*(\w+)$',
            r'to\s*(\d+\.?\d*)\s*(\w+)$',
            r'→\s*(\d+\.?\d*)\s*(\w+)$'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, notes)
            if match:
                amount, symbol = match.groups()
                return {
                    'symbol': symbol.upper(),
                    'amount': Decimal(str(amount))
                }
        
        logging.warning(f"Could not extract received asset from notes: {notes}")
        return None
        
    except Exception as e:
        log_error(f"Error parsing conversion notes: {notes}")
        log_error(str(e))
        return None

def validate_input_file(df: pd.DataFrame) -> None:
    # Validate input file structure and contents

    required_columns = {
        'ID', 'Timestamp', 'Source', 'Type', 'Asset', 
        'Amount', 'Total USD', 'Fee', 'Spot Price', 'Notes'
    }
    
    # Check for required columns
    missing_columns = required_columns - set(df.columns)
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # Validate data types
    if not pd.api.types.is_datetime64_any_dtype(df['Timestamp']):
        raise ValueError("Timestamp column must contain datetime values")
    
    # Validate numeric columns
    numeric_columns = ['Amount', 'Total USD', 'Fee', 'Spot Price']
    for col in numeric_columns:
        if not pd.to_numeric(df[col], errors='coerce').notna().all():
            raise ValueError(f"Column {col} must contain numeric values")
    
    # Validate required fields
    if df['ID'].isna().any():
        raise ValueError("Transaction IDs cannot be null")
    if df['Asset'].isna().any():
        raise ValueError("Asset symbol cannot be null")
    if df['Type'].isna().any():
        raise ValueError("Transaction type cannot be null")

def strip_trailing_zeros(number: Union[Decimal, float, str]) -> str:
    #Format number as string, removing trailing zeros but preserving small values
    if number is None:
        return '0'
        
    # Convert to Decimal for consistent handling
    if not isinstance(number, Decimal):
        number = Decimal(str(number))
    
    # Format with enough decimal places to handle small numbers
    formatted = f"{number:.12f}"
    
    # Remove trailing zeros after decimal point, but keep decimal if whole number
    if '.' in formatted:
        formatted = formatted.rstrip('0').rstrip('.')
        if formatted == '0' and number != 0:
            # For very small numbers, format with fixed decimal places
            formatted = f"{number:.8f}"
    
    return formatted

def normalize_usd(value: Union[float, Decimal, str]) -> Decimal:
    """Normalize USD amount to 2 decimal places"""
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    return value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def calculate_gains(transactions_df: pd.DataFrame, accounting_method: AccountingMethod = AccountingMethod.HIFO) -> pd.DataFrame:
    # Calculate capital gains/losses for all cryptocurrency transactions.
    # Uses specified accounting method to determine which lots to sell first.
    # 
    # Parameters:
    # - transactions_df: DataFrame containing all transactions
    # - accounting_method: Which method to use (FIFO, LIFO, HIFO, or LOFO)
    # 
    # Returns DataFrame with calculated gains/losses for each sale

    try:
        # Make a copy and convert Amount column to Decimal
        df = transactions_df.copy()
        
        # Ensure Lot ID is string type
        df['Lot ID'] = df['Lot ID'].astype(str)
        
        # Convert numeric columns to Decimal, preserving ALL significant digits
        # Don't normalize/round crypto amounts to maintain maximum precision
        df['Amount'] = df['Amount'].apply(lambda x: Decimal(str(x)) if pd.notna(x) else Decimal('0'))
        
        # For USD values, preserve full precision during calculations
        df['Total USD'] = df['Total USD'].apply(lambda x: Decimal(str(x)) if pd.notna(x) else Decimal('0'))
        df['Subtotal'] = df['Subtotal'].apply(lambda x: Decimal(str(x)) if pd.notna(x) else Decimal('0'))
        
        # Pre-calculate and validate total amounts for each asset
        asset_totals = defaultdict(Decimal)
        for idx, row in df.iterrows():
            # Skip USD deposits/withdrawals, Exchange Withdrawals, Pro Withdrawals, and Send transactions
            if (row['Asset'] == 'USD' and row['Type'].lower() in ['deposit', 'withdrawal']) or \
               'exchange withdrawal' in row['Type'].lower() or \
               'pro withdrawal' in row['Type'].lower() or \
               'send' in row['Type'].lower():
                continue
                
            asset_totals[row['Asset']] += df['Amount'][idx]
        
        # Collect negative balance warnings
        validation_errors = []
        for asset, total in asset_totals.items():
            if total < 0 and asset != 'USD':
                validation_errors.append({
                    'timestamp': None,
                    'asset': asset,
                    'error': f'Negative total balance: {total}',
                    'transaction': None
                })
        
        sales = []
        lots = defaultdict(list)
        
        # Process transactions chronologically
        for idx, row in df.sort_values(['Timestamp', 'Type']).iterrows():
            asset = row['Asset']
            amount = df['Amount'][idx]
            
            if amount > 0:  # Buy/receive
                if row['Lot ID']:  # Only track if it has a lot ID
                    total_cost = abs(df['Total USD'][idx])  # Use absolute value for cost basis
                    basis_per_unit = (total_cost / amount) if amount != 0 else Decimal('0')
                    lots[asset].append({
                        'lot_id': row['Lot ID'],
                        'remaining': amount,
                        'original': amount,
                        'timestamp': row['Timestamp'],
                        'basis_per_unit': basis_per_unit
                    })
            elif amount < 0:  # Sell/send
                # Skip certain transaction types for lot validation
                if ('withdrawal' in row['Type'].lower() or
                    'send' in row['Type'].lower()):
                    continue

                remaining_sale = abs(amount)
                used_lots = []
                
                # Sort lots once based on accounting method
                if accounting_method == AccountingMethod.FIFO:
                    lots[asset].sort(key=lambda x: (x['timestamp'], Decimal(str(x['basis_per_unit']))))
                elif accounting_method == AccountingMethod.LIFO:
                    lots[asset].sort(key=lambda x: (-x['timestamp'].timestamp(), Decimal(str(x['basis_per_unit']))))
                elif accounting_method == AccountingMethod.HIFO:
                    lots[asset].sort(key=lambda x: (-Decimal(str(x['basis_per_unit'])), x['timestamp']))
                elif accounting_method == AccountingMethod.LOFO:
                    lots[asset].sort(key=lambda x: (Decimal(str(x['basis_per_unit'])), x['timestamp']))

                # Process lots in the sorted order
                for lot in lots[asset]:
                    if remaining_sale <= 0:
                        break
                    
                    if lot['remaining'] > 0:
                        used_amount = min(lot['remaining'], remaining_sale)
                        remaining_sale = remaining_sale - used_amount
                        lot['remaining'] = lot['remaining'] - used_amount
                        
                        used_lots.append({
                            'lot_id': lot['lot_id'],
                            'amount_sold': used_amount,
                            'basis_per_unit': lot['basis_per_unit'],
                            'purchase_date': lot['timestamp']
                        })
                
                if used_lots:
                    # Use Subtotal for sale price (before fees)
                    sale_price = abs(row['Subtotal'] / amount) if amount != 0 else Decimal('0')
                    
                    # Add details for each lot used
                    for lot in used_lots:
                        sale_detail = {
                            'Sale Date': row['Timestamp'],
                            'Purchase Date': lot['purchase_date'],
                            'Asset': asset,
                            'Lot ID': str(lot['lot_id']),
                            'Amount Sold': lot['amount_sold'],
                            'Cost Basis Per Unit': lot['basis_per_unit'],
                            'Sale Price Per Unit': sale_price,
                            'Sale ID': str(row['ID'])
                        }
                        sales.append(sale_detail)
        
        # Create sales DataFrame with calculated gains
        if sales:
            sales_df = pd.DataFrame(sales)
            
            # Calculate final values, maintaining precision until the last step
            # Only normalize USD amounts at the very end
            sales_df['Cost Basis'] = sales_df.apply(
                lambda x: normalize_usd(
                    Decimal(str(x['Amount Sold'])) * Decimal(str(x['Cost Basis Per Unit']))
                ),
                axis=1
            )
            sales_df['Proceeds'] = sales_df.apply(
                lambda x: normalize_usd(
                    Decimal(str(x['Amount Sold'])) * Decimal(str(x['Sale Price Per Unit']))
                ),
                axis=1
            )
            # Calculate Gain/Loss after normalizing the components
            sales_df['Gain/Loss'] = sales_df['Proceeds'] - sales_df['Cost Basis']
            
            return sales_df, validation_errors
        
        return pd.DataFrame(), validation_errors  # Return empty DataFrame if no sales
        
    except Exception as e:
        print(f"Error calculating gains: {str(e)}")
        raise

def get_historical_price(asset: str, timestamp: datetime, current: int = None, total: int = None) -> float:
    # Get the historical price of a cryptocurrency at a specific point in time.
    # This function is critical for calculating cost basis and gains.
    #
    # The process works like this:
    # 1. First checks our local cache file (price_cache.json)
    # 2. If not found in cache, queries the CryptoCompare API:
    #    - With API key: Gets minute-by-minute prices (last 7 days only)
    #    - Without API key: Gets daily closing prices
    # 3. Handles special cases like the LUNA -> LUNC transition & CGLD -> CELO
    # 4. Saves successful price lookups to cache for future use
    # 5. If price lookup fails, caches the failure for 12 hours
    #
    # Parameters:
    # - asset: Cryptocurrency symbol to look up (e.g., 'BTC', 'ETH')
    # - timestamp: Exact date and time we need the price for
    # - current/total: Optional progress indicators for batch processing
    #
    # Returns:
    # - Historical price in USD, or 0 if price not found

    global price_cache
    
    # Format date for display in logs and messages
    date_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
    count_str = f"{current}/{total}: " if current and total else ""
    
    # Handle the LUNA & LUNA2 transition when LUNA became LUNC, and LUNA2 became the new LUNA
    display_name = asset
    if asset == 'LUNA':
        if timestamp >= LUNA_TRANSITION_DATE:
            asset = 'LUNC'  # Use LUNC for price lookups after transition
            display_name = 'LUNA (AKA: LUNC)'
    elif asset == 'LUNA2':
        if timestamp >= LUNA_TRANSITION_DATE:
            asset = 'LUNA'  # Use new LUNA for price lookups after transition
            display_name = 'LUNA2 (AKA: LUNA)'
    elif asset == 'NU':
        if timestamp >= NU_TRANSITION_DATE:
            asset = 'T'
            display_name = 'NU (AKA: T) **** NuCypher merged into [T]hreshold Network **** 1 NU = 3.26 T'
    elif asset == 'CGLD':
        asset = 'CELO'
        display_name = 'CGLD (AKA: CELO)'
       
    # Determine if we have an API KEY to use minute-level price data
    use_minute_data = CRYPTOCOMPARE_API_KEY and CRYPTOCOMPARE_API_KEY != 'YOUR-API-KEY-HERE'
    
    # Round the timestamp based on available data granularity
    # Minute-level for recent data with API key, daily otherwise
    if use_minute_data:
        rounded_timestamp = timestamp.replace(second=0, microsecond=0)
        cache_key = f"{asset}_{rounded_timestamp.strftime('%Y-%m-%d_%H-%M')}"
    else:
        rounded_timestamp = timestamp.replace(hour=0, minute=0, second=0, microsecond=0)
        cache_key = f"{asset}_{rounded_timestamp.strftime('%Y-%m-%d')}"
    
    # Key for tracking failed price lookups to avoid repeated API calls
    no_price_key = f"no_price_{asset}_{rounded_timestamp.strftime('%Y-%m-%d')}"
    
    # Check cache first to avoid unnecessary API calls
    if cache_key in price_cache:
        cached_data = price_cache[cache_key]
        if isinstance(cached_data, dict) and 'price' in cached_data:
            # If price found and cache is fresh (less than 12 hours old)
            if cached_data['timestamp'] and (int(datetime.now().timestamp()) - cached_data['timestamp']) < 43200:
                price_str = strip_trailing_zeros(cached_data['price'])
                print(f"Processing {count_str}{display_name} at {date_str} - CACHED SUCCESS - Price: ${price_str}")
                return float(cached_data['price'])
    
    # Check no-price cache to avoid repeated API calls for known missing prices
    if no_price_key in price_cache:
        cached_data = price_cache[no_price_key]
        if isinstance(cached_data, dict) and 'timestamp' in cached_data:
            cache_age = datetime.now() - datetime.fromtimestamp(cached_data['timestamp'])
            if cache_age.total_seconds() < 43200:  # No-price cache valid for 12 hours
                print(f"Processing {count_str}{display_name} at {date_str} - FAILED - Price: $0.00")
                return 0
        # If cache entry is expired, remove it so we can try fetching again
        del price_cache[no_price_key]
    
    try:
        # Convert timestamp to UNIX timestamp for API call
        unix_time = int(rounded_timestamp.timestamp())
        
        params = {
            'fsym': asset,
            'tsym': 'USD',
            'limit': 1,
            'toTs': unix_time
        }

        # Try minute data first if available
        if use_minute_data and (datetime.now() - rounded_timestamp).days <= 7:
            url = 'https://min-api.cryptocompare.com/data/v2/histominute'
            params['api_key'] = CRYPTOCOMPARE_API_KEY
        else:
            # Fall back to daily data
            url = 'https://min-api.cryptocompare.com/data/v2/histoday'
        
        headers = {}
        
        # Add delay to avoid rate limiting
        time.sleep(0.25)
        
        # Make API request with timeout
        response = requests.get(url, params=params, headers=headers, timeout=10)  # Add 10 second timeout
        data = response.json()
        
        if response.status_code != 200:
            error_msg = f"API Error ({response.status_code}): {data.get('Message', 'Unknown error')}"
            log_error(error_msg)
            return 0
        
        # Try to get price from response
        if data.get('Response') == 'Success' and data.get('Data', {}).get('Data'):
            close_price = data['Data']['Data'][0].get('close', 0)
            
            # If no price found and using daily data, try previous day
            if close_price == 0 and not use_minute_data:
                yesterday = rounded_timestamp - timedelta(days=1)
                unix_time = int(yesterday.timestamp())
                params['toTs'] = unix_time
                
                time.sleep(0.25)  # Add delay for second request
                response = requests.get(url, params=params, headers=headers)
                data = response.json()
                
                if response.status_code == 200 and data.get('Response') == 'Success' and data.get('Data', {}).get('Data'):
                    close_price = data['Data']['Data'][0].get('close', 0)
            
            # When price is found:
            if close_price > 0:
                price_str = strip_trailing_zeros(close_price)
                print(f"Processing {count_str}{display_name} at {date_str} - FETCH SUCCESS - Price: ${price_str}")
                price_cache[cache_key] = {
                    'price': float(close_price),
                    'timestamp': int(datetime.now().timestamp())
                }
                save_price_cache(price_cache)
                return float(close_price)
        
        # If we get here, no price was found
        print(f"Processing {count_str}{display_name} at {date_str} - FAILED - Price: $0.00")
        
        # Cache the no-price result
        price_cache[no_price_key] = {
            'timestamp': int(datetime.now().timestamp())
        }
        save_price_cache(price_cache)
        return 0
        
    except requests.exceptions.Timeout:
        print(f"Processing {count_str}{display_name} at {date_str} - TIMEOUT - Price: $0.00")
        return 0
    except requests.exceptions.RequestException as e:
        logging.error(f"Network error fetching price for {display_name}: {str(e)}")
        print(f"Processing {count_str}{display_name} at {date_str} - FAILED - Price: $0.00")
        return 0
    except Exception as e:
        logging.error(f"Unexpected error getting price for {display_name}: {str(e)}")
        print(f"Processing {count_str}{display_name} at {date_str} - FAILED - Price: $0.00")
        return 0

def verify_final_calculations(yearly_data: Dict, processor: TransactionProcessor) -> None:
    # Verify all calculations are consistent
    # Check for mismatches in proceeds and amounts
    # Store any discrepancies for later reporting
    
    # Original verification logic with more detailed output
    for year, data in yearly_data.items():
        for symbol, asset_data in data.items():
            # Verify short-term calculations
            for sale in asset_data['short_term_sales']:
                total_lot_amounts = sum(detail.amount_sold for detail in sale.lot_details)
                if abs(total_lot_amounts - abs(sale.amount)) > Decimal('0.00000001'):
                    print(f"Amount mismatch in short-term sale:")
                    print(f"Total from lots: {total_lot_amounts}")
                    print(f"Sale amount: {abs(sale.amount)}")
                    print(f"Difference: {abs(total_lot_amounts - abs(sale.amount))}")
                    raise ValueError(f"Amount mismatch in {symbol} short-term sale")

            # Verify long-term calculations with higher tolerance
            for sale in asset_data['long_term_sales']:
                total_proceeds = sum(detail.proceeds for detail in sale.lot_details)
                proceeds_difference = abs(total_proceeds - sale.proceeds)
                
                # Use a more generous tolerance for proceeds verification
                tolerance = Decimal('0.02')  # 2 cents tolerance
                
                if proceeds_difference > tolerance:
                    # Store the mismatch details in the processor for later display
                    processor.add_proceeds_mismatch(symbol, sale.timestamp, total_proceeds, sale.proceeds)
                    # Also store the lot details
                    processor.add_mismatch_details(symbol, {
                        'timestamp': sale.timestamp,
                        'amount': sale.amount,
                        'expected_proceeds': total_proceeds,
                        'actual_proceeds': sale.proceeds,
                        'difference': proceeds_difference,
                        'lot_details': [
                            {
                                'amount': detail.amount_sold,
                                'proceeds': detail.proceeds,
                                'cost_basis': detail.cost_basis
                            } for detail in sale.lot_details
                        ]
                    })

def check_accounting_method_consistency(df: pd.DataFrame) -> None:
    # Check if accounting method changes across years

    years = df['Timestamp'].dt.year.unique()
    if len(years) > 1:
        print("\nWARNING: Processing multiple years with same accounting method.")
        print("IRS regulations may require consistent use of accounting method across years.")
        input("Press Enter to continue or Ctrl+C to abort...")

def validate_accounting_method_consistency(df: pd.DataFrame, method: AccountingMethod) -> None:
    # Ensure accounting method is valid for the given transactions
    # Check if switching methods from previous years

    previous_method_file = Path('previous_accounting_method.txt')
    if previous_method_file.exists():
        with open(previous_method_file, 'r') as f:
            previous_method = f.read().strip()
            if previous_method != method.name:
                raise ValueError(
                    f"Accounting method changed from {previous_method} to {method.name}. "
                    "IRS regulations require consistent use of accounting methods."
                )
    
    # Save current method
    with open(previous_method_file, 'w') as f:
        f.write(method.name)

def compare_all_method_outputs():
    # Compare outputs from all accounting methods and highlight differences
    # Matches rows by Asset column when comparing across methods, except for
    # Transfers and Sale Details sheets which compare row by row
    
    # Define USD value columns (case-insensitive)
    usd_columns = {'cost', 'basis', 'gain', 'loss', 'price', 'proceeds'}
    
    methods = [m.name for m in AccountingMethod]
    workbooks = {
        method: openpyxl.load_workbook(f'ALL-crypto-profit-and-loss-{method}.xlsx')
        for method in methods
    }
    
    # Compare each sheet in the workbooks
    sheet_names = workbooks[methods[0]].sheetnames
    for sheet_name in sheet_names:
        # Get all values for each cell position across methods
        for method, wb in workbooks.items():
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            
            # Get column headers to identify USD columns
            headers = [cell.value.lower() if cell.value else '' for cell in next(ws.rows)]
            
            # Handle Transfers and Sale Details sheets with original row-by-row comparison
            if sheet_name in ['Transfers', 'Sale Details']:
                max_rows = max(ws.max_row for wb in workbooks.values() for ws in [wb[sheet_name]] if sheet_name in wb.sheetnames)
                max_cols = max(ws.max_column for wb in workbooks.values() for ws in [wb[sheet_name]] if sheet_name in wb.sheetnames)
                
                # Compare each cell row by row
                for row in range(2, max_rows + 1):  # Skip header row
                    for col in range(1, max_cols + 1):
                        values = set()
                        header = headers[col-1]
                        
                        # Collect values from all methods
                        for m in methods:
                            if sheet_name in workbooks[m].sheetnames:
                                cell = workbooks[m][sheet_name].cell(row=row, column=col)
                                if cell.value is not None:
                                    # For USD columns, round to 2 decimal places
                                    if any(usd_term in str(header).lower() for usd_term in usd_columns):
                                        try:
                                            val = str(cell.value).replace('$', '').replace(',', '')
                                            values.add(round(float(val), 2))
                                        except:
                                            values.add(str(cell.value))
                                    else:
                                        values.add(str(cell.value))
                        
                        # If values differ, highlight cells in all workbooks
                        if len(values) > 1:
                            yellow_fill = openpyxl.styles.PatternFill(
                                start_color='FFFF99',
                                end_color='FFFF99',
                                fill_type='solid'
                            )
                            
                            for m in methods:
                                if sheet_name in workbooks[m].sheetnames:
                                    cell = workbooks[m][sheet_name].cell(row=row, column=col)
                                    cell.fill = yellow_fill
                continue  # Move to next sheet
            
            # For all other sheets, compare by Asset
            # Find the Asset column index (usually column A)
            header_row = next(ws.rows)
            asset_col_idx = None
            for idx, cell in enumerate(header_row, 1):
                if cell.value == 'Asset':
                    asset_col_idx = idx
                    break
            
            if asset_col_idx is None:
                continue
                
            # Create a mapping of Asset to row number for each method
            asset_row_maps = {}
            for m in methods:
                if sheet_name in workbooks[m].sheetnames:
                    curr_ws = workbooks[m][sheet_name]
                    asset_row_maps[m] = {}
                    for row in range(2, curr_ws.max_row + 1):  # Skip header row
                        asset_cell = curr_ws.cell(row=row, column=asset_col_idx)
                        if asset_cell.value:
                            asset_row_maps[m][asset_cell.value] = row
            
            # Get all unique assets across all methods
            all_assets = set()
            for row_map in asset_row_maps.values():
                all_assets.update(row_map.keys())
            
            # Compare each asset's data across methods
            for asset in all_assets:
                # Get the maximum number of columns across all methods
                max_cols = max(
                    workbooks[m][sheet_name].max_column 
                    for m in methods 
                    if sheet_name in workbooks[m].sheetnames
                )
                
                # Compare each column for this asset
                for col in range(1, max_cols + 1):
                    values = set()
                    header = headers[col-1]
                    
                    # Collect values from all methods for this asset and column
                    for m in methods:
                        if sheet_name in workbooks[m].sheetnames:
                            if asset in asset_row_maps[m]:
                                row = asset_row_maps[m][asset]
                                cell = workbooks[m][sheet_name].cell(row=row, column=col)
                                if cell.value is not None:
                                    # For USD columns, round to 2 decimal places
                                    if any(usd_term in str(header).lower() for usd_term in usd_columns):
                                        try:
                                            val = str(cell.value).replace('$', '').replace(',', '')
                                            values.add(round(float(val), 2))
                                        except:
                                            values.add(str(cell.value))
                                    else:
                                        values.add(str(cell.value))
                    
                    # If values differ, highlight cells in all workbooks
                    if len(values) > 1:
                        yellow_fill = openpyxl.styles.PatternFill(
                            start_color='FFFF99',
                            end_color='FFFF99',
                            fill_type='solid'
                        )

                        for m in methods:
                            if sheet_name in workbooks[m].sheetnames:
                                if asset in asset_row_maps[m]:
                                    row = asset_row_maps[m][asset]
                                    cell = workbooks[m][sheet_name].cell(row=row, column=col)
                                    cell.fill = yellow_fill
    
    # Save all workbooks
    for method, wb in workbooks.items():
        wb.save(f'ALL-crypto-profit-and-loss-{method}.xlsx')

def format_holdings_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    # Format the holdings DataFrame with proper column order and formatting
    # Handles column ordering, data type conversion, and column renaming
    # 
    # Processes these types of columns:
    # - Amount columns (with variable decimal places)
    # - Price columns (standardized to 8 decimals)
    # - USD columns (standardized to 2 decimals)
    # - Date columns
    # - Percentage columns

    today = datetime.now()
    spot_price_key = f'spot_price_{today.strftime("%Y-%m-%d")}'
    total_usd_key = f'total_usd_{today.strftime("%Y-%m-%d")}'
    
    # Convert numeric columns to appropriate types
    numeric_columns = [
        'total_amount',
        'average_cost_basis',
        'total_cost_basis',
        'total_fees',
        spot_price_key,
        total_usd_key,
        'unrealized_pl',
        'unrealized_pl_pct',
        'lot_count',
        'days_held',
        'avg_hold_time',
        'long_term_amount',
        'short_term_amount',
        'lowest_price',
        'highest_price'
    ]
    
    # Convert each numeric column
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace('$', '').str.replace(',', ''), errors='coerce')
    
    # Get the remaining lots for each asset from the lots data
    lots_remaining = defaultdict(list)
    if 'lots' in df.columns:  # Check if lots data exists
        for idx, row in df.iterrows():
            if row['lots']:  # If there are lots
                for lot in row['lots']:
                    if lot.remaining > 0:
                        lots_remaining[row['symbol']].append(
                            f"Lot {lot.lot_id}: {lot.remaining}"
                        )
    
    # Reorder columns
    column_order = [
        'symbol',                   # Asset
        'total_amount',             # Total Amount
        'average_cost_basis',       # Average Cost Basis
        'total_cost_basis',         # Total Cost Basis
        'total_fees',               # Total Fees Paid
        spot_price_key,             # Current spot price
        total_usd_key,              # Current total USD value
        'unrealized_pl',            # Unrealized Profit/Loss (USD)
        'unrealized_pl_pct',        # Unrealized Profit/Loss (%)
        'exchange_distribution',    # Current Exchange Distribution
        'last_tx_date',             # Last Tax Lot Date
        'days_held',                # Days Held (from first Tx)
        'avg_hold_time',            # Average Hold Time
        'long_term_amount',         # Long-term Holdings Amount
        'short_term_amount',        # Short-term Holdings Amount
        'lowest_price',             # Lowest Purchase Price
        'highest_price',            # Highest Purchase Price
        'lots_remaining',           # Lots Remaining
        'lot_count'                 # Number of Lots
        ]
    
    df = df[column_order]
    
    # Rename columns for display
    df.columns = [
        'Asset',
        'Total Amount',
        'Average Cost Basis',
        'Total Cost Basis',
        'Total Fees Paid',
        f'Spot Price on {today.strftime("%Y-%m-%d")}',
        f'Total USD on {today.strftime("%Y-%m-%d")}',
        'Unrealized Profit/Loss (USD)',
        'Unrealized Profit/Loss (%)',
        'Current Exchange Distribution',
        'Last Tax Lot Date',
        'Days Held (from first Tx)',
        'Average Hold Time (days)',
        'Long-term Holdings Amount',
        'Short-term Holdings Amount',
        'Lowest Purchase Price',
        'Highest Purchase Price',
        'Lots Remaining',
        'Number of Lots'
    ]
    

    return df

def format_excel_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame = None, sheet_name: str = '') -> None:
    # Comprehensive worksheet formatting function that handles all formatting needs
    # - Adds filters to all columns
    # - Applies consistent header styling (light blue background, bold text)
    # - Calculates and sets optimal column widths based on content
    # - Handles special formatting
    
    # Add filters
    if df is not None:
        worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df) + 1}"
    
    # Define styles
    header_fill = openpyxl.styles.PatternFill(
        start_color='CCE5FF',  # Light blue
        end_color='CCE5FF',
        fill_type='solid'
    )
    header_font = openpyxl.styles.Font(bold=True)
    center_align = openpyxl.styles.Alignment(horizontal='center')
    
    # First pass: Apply basic header formatting and calculate content lengths
    column_content_lengths = {}
    proceeds_col_letter = None
    cost_basis_col_letter = None
    proceeds_width = 0
    cost_basis_width = 0
    
    for column in worksheet.columns:
        col_letter = column[0].column_letter
        header_cell = column[0]
        header_value = str(header_cell.value)
        
        # Apply basic header formatting
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = center_align  # Default to center
        
        # Calculate max content length (excluding header)
        max_content_length = 0
        for cell in column[1:]:  # Skip header
            try:
                if cell.value:
                    max_content_length = max(max_content_length, len(str(cell.value)))
            except:
                pass
        column_content_lengths[col_letter] = max_content_length
        
        # Track Proceeds and Cost Basis columns for this sheet
        if header_value == 'Proceeds':
            proceeds_col_letter = col_letter
            proceeds_width = max(len(header_value) + 4, max_content_length + 2)
        elif header_value == 'Cost Basis':
            cost_basis_col_letter = col_letter
            cost_basis_width = max(len(header_value) + 4, max_content_length + 2)
    
    # Second pass: Handle column widths, alignments, and number formats
    for column in worksheet.columns:
        col_letter = column[0].column_letter
        header_cell = column[0]
        header_value = str(header_cell.value)
        header_length = len(header_value)
        max_content_length = column_content_lengths[col_letter]
        
        # Calculate base column width
        content_width = max(max_content_length + 2, 10)  # Minimum 10 characters
        header_width = header_length + 4  # Add space for filter dropdown
        
        # Add extra width for specific columns
        extra_width = 0
        if sheet_name.startswith('Year_'):
            if 'Total Amount' in header_value:
                extra_width = 4
        elif sheet_name == 'Current Holdings':
            if 'Total USD' in header_value or 'Holdings Amount' in header_value:
                extra_width = 2
            elif 'Total Amount' in header_value:
                extra_width = 4
        elif sheet_name == 'Transfers':
            if 'From Exchange' in header_value:
                extra_width = 2
        
        # Special handling for Proceeds and Cost Basis columns in this sheet
        if proceeds_col_letter and cost_basis_col_letter:
            max_width = max(proceeds_width, cost_basis_width)
            if col_letter in (proceeds_col_letter, cost_basis_col_letter):
                final_width = max_width
            else:
                final_width = min(max(content_width, header_width + extra_width, 10), 100)
        else:
            final_width = min(max(content_width, header_width + extra_width, 10), 100)
        
        # Set column width
        worksheet.column_dimensions[col_letter] = openpyxl.worksheet.dimensions.ColumnDimension(
            worksheet, 
            index=col_letter, 
            width=final_width,
            bestFit=False
        )
        
        # Apply number formatting to data cells
        header_value_lower = header_value.lower()
        for cell in column[1:]:  # Skip header
            # Skip percentage columns
            if '%' in header_value_lower or 'percent' in header_value_lower:
                cell.number_format = '0.00%;[Red]-0.00%'  # Percentage with red negatives
            # Special handling for any column containing 'loss' or 'total fees'
            elif 'loss' in header_value_lower or 'total fees' in header_value_lower:
                cell.number_format = '_($* #,##0.00_);[Red]_($* -#,##0.00_)'  # Left-aligned $ with red minus
            # Regular handling for other monetary columns
            elif any(x in header_value_lower for x in ['price', 'cost', 'basis', 'proceeds', 'usd', 'fees']):
                cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[44]  # Standard currency format
            elif 'amount' in header_value_lower:
                if cell.value is not None:
                    # Count significant decimal places in the actual value
                    str_val = str(cell.value)
                    if '.' in str_val:
                        decimal_part = str_val.split('.')[1]
                        # Find the last non-zero digit position
                        last_nonzero = len(decimal_part)
                        for i in range(len(decimal_part) - 1, -1, -1):
                            if decimal_part[i] != '0':
                                last_nonzero = i + 1
                                break
                        decimals = last_nonzero
                        if decimals > 0:
                            cell.number_format = f'#,##0.{"0" * decimals}'
                        else:
                            cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0'
            elif 'date range' in header_value_lower:
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
            elif 'timestamp' in header_value_lower:
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            elif 'date' in header_value_lower:
                cell.number_format = 'YYYY-MM-DD'  # Date

class InsufficientLotsError(Exception):
    # Custom exception for insufficient lots
    pass

def log_error(message: str, print_to_screen: bool = True) -> None:
    # Helper function to handle error logging
    logging.getLogger().error(message)
    if print_to_screen:
        print(f"ERROR: {message}")

def report_issues(all_issues: List[Dict]) -> None:
    """Report any issues found during processing"""
    if all_issues:
        print("\n\n################################################")
        print("************** ISSUES DETECTED *****************")
        print("################################################")
        print("")

        for method_issues in all_issues:
            print("-" * (len(method_issues['method']) + 1))            
            print(f"{method_issues['method']}:")
            print("-" * (len(method_issues['method']) + 1))
            
            if method_issues['errors']:
                for error in method_issues['errors']:
                    tx = error.get('transaction', {})
                    
                    # Parse the error message to extract Need/Have values
                    error_msg = error['error']
                    if 'Insufficient lots' in error_msg:
                        print(f"ERROR: {error['error']}:")
                        print(f"ERROR: Need: {tx.get('Need')}, Have: {tx.get('Have')}")
                        print(f"ERROR: Error processing transaction: {tx.get('ID')} at {error['timestamp']}")
                        print(f"ERROR: Error details: {error['error']}")
                        print("")
            
            if method_issues['proceeds_mismatches']:
                print("\nProceeds mismatches found:")
                for mismatch in method_issues['proceeds_mismatches']:
                    print(f"\n- {mismatch['symbol']}:")
                    print(f"  Time: {mismatch['timestamp']}")
                    print(f"  Difference: ${mismatch['difference']:.2f}")
                
                for detail in method_issues['mismatch_details']:
                    print(f"\nDetails for {detail['symbol']} at {detail['timestamp']}:")
                    print(f"  Amount: {detail['amount']}")
                    print(f"  Expected: ${detail['expected_proceeds']:.2f}")
                    print(f"  Actual: ${detail['actual_proceeds']:.2f}")
                    print(f"  Difference: ${detail['difference']:.2f}")
                    
                    if detail.get('lot_details'):
                        print("\n  Lots used:")
                        for lot in detail['lot_details']:
                            print(f"    Amount: {lot['amount']}")
                            print(f"    Proceeds: ${lot['proceeds']:.2f}")
                            print(f"    Cost basis: ${lot['cost_basis']:.2f}")
                            print("    ---")
        
        print("\n\n*************************************************************")
        print("See full details in the 'ALL-crypto-profit-and-loss-errors.log' file")
        print("*************************************************************\n")
    else:
        print("\n\n**********************************")
        print("No issues found during processing.")
        print("**********************************\n")

def main():
    try:
        print("\n~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("Select accounting method:")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~")
        
        print("ALL:  Generate reports for *ALL* methods")
        for method in AccountingMethod:
            print(f"{method.name}: {method.value}")
        
        while True:
            method_input = input("\n\nEnter method (ALL/FIFO/LIFO/HIFO/LOFO) [default: ALL]: ").upper()
            if not method_input or method_input == 'ALL':  # Handle both empty input and 'ALL'
                methods = list(AccountingMethod)  # Default is ALL methods
                method_input = 'ALL'  # Set to ALL for later comparison check
                break
            try:
                methods = [AccountingMethod[method_input]]
                break
            except KeyError:
                print("Invalid method. Please try again.")

        # Read and validate input file
        df = pd.read_excel('ALL-MASTER-crypto-transactions.xlsx')
        validate_input_file(df)

        all_issues = []  # Track all issues across methods
        # last_processor = None  # Keep track of the last processor for caching info

        # Process each selected method
        for method in methods:
            print("\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            print(f"Processing using {method.value}...")
            print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")
            
            # Add a separator line in the log file for this method
            logging.info(f"\n{'='*80}\nProcessing with {method.value}\n{'='*80}")
            
            # Initialize processor with selected method
            processor = TransactionProcessor(accounting_method=method)
            # last_processor = processor  # Store for later use
            
            # Process all transactions
            processor.process_all_transactions(df)
            
            # Calculate gains and generate report
            calculator = GainsCalculator(processor)
            yearly_data = calculator.calculate_yearly_summary()
            
            # Verify all calculations before generating report
            verify_final_calculations(yearly_data, processor)
            
            # Store any issues found
            if processor.all_errors or processor.proceeds_mismatches:
                method_issues = {
                    'method': method.value,
                    'errors': processor.all_errors.copy(),
                    'proceeds_mismatches': processor.proceeds_mismatches.copy(),
                    'mismatch_details': processor.get_mismatch_details()
                }
                all_issues.append(method_issues)
            
            # Generate report with method-specific suffix
            suffix = f'-{method.name}'
            generate_excel_report(yearly_data, processor, suffix)
            
            print(f"Calculations verified successfully for:  {method.value}")
            print("                                         *********************")
        
        print("\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("*ALL* reports generated successfully!!")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        
        # If ALL methods were processed, compare and highlight differences
        if method_input == 'ALL':
            print("\nComparing outputs across all methods...")
            compare_all_method_outputs()
            print("\nComparison complete -  Differences highlighted in YELLOW")
            print("                                                  ******")

        # Report any issues found during processing
        report_issues(all_issues)

    except Exception as e:
        log_error(f"Error in main execution: {str(e)}")
        raise

if __name__ == "__main__":
    main()